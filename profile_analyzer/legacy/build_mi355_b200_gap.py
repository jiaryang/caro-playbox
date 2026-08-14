import json, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ATOM = r"C:\Workspace\qwen\Atom_Qwen3.5-397B-A17B-MXFP4_ts_20260728_101711_991.pt.trace.json"
SGL  = r"C:\Workspace\qwen\profile_qwen_mi355_8k1k_c128_1784806436.739748\qwen_mi355_8k1k_c128-1784806436.7420862-TP-0-DECODE.trace.json"
B200 = r"C:\Workspace\qwen\profile_qwen_b200_8k1k_c128_1784800637.471206\qwen_b200_8k1k_c128-1784800637.475314-TP-0-DECODE.trace.json"
OUT  = r"C:\Workspace\qwen\qwen_decode_gap_mi355_vs_b200.xlsx"

def cat_of(name):
    n = name.lower()
    is_tri = n.startswith("triton") or "fused" in n
    if ("cross_device_reduce" in n or "nccl" in n or "rccl" in n or "all_gather" in n or "allgather" in n
        or "reduce_scatter" in n or "quickreduce" in n or "custom_ar" in n or "lamport" in n
        or (("all_reduce" in n or "allreduce" in n) and not is_tri)):
        return "communication"
    if ("moe" in n or "topkgating" in n or "expert" in n or "grouped_gemm" in n or "group_gemm" in n
        or "routing" in n or ("bmm" in n and ("e2m1" in n or "e4m3" in n or "nvfp4" in n))):
        return "moe"
    if any(k in n for k in ["gated_delta","gdn","causal_conv","recurrent_gated","chunk_gated","split_chunk","chunk_fwd","recompute_w_u"]):
        return "mixer:linear(gdn)"
    if any(k in n for k in ["fmha","flash_attn","flashinfer","_attn_","paged_attn","paged_attention","mha","attention","_attn"]):
        return "mixer:full-attn"
    if (n.startswith("cijk") or "gemm" in n or "_mm_" in n or "matmul" in n or "hgemm" in n or "nvjet" in n
        or "cutlass" in n or "cublas" in n or "splitkreduce" in n):
        return "gemm"
    if "rsqrt" in n or "rmsnorm" in n or "layernorm" in n or "qk_norm" in n or "layer_norm" in n:
        return "normalization"
    if "quant" in n or "fp4" in n or "fp8" in n or "preshuffle" in n or "scaled" in n or "absmax" in n:
        return "quantization"
    if any(k in n for k in ["act_and_mul","silu","gelu","sigmoid","softmax"]):
        return "activation"
    if any(k in n for k in ["copybuffer","memcpy","memset","__amd_rocclr","dtod","htod"]):
        return "memcpy"
    if any(k in n for k in ["elementwise","embedding","rope","rotary","index","cat","slice","add","mul","fill","_to_copy","direct_copy"]):
        return "elementwise"
    return "other"

KCAT = {"kernel","gpu_memcpy","gpu_memset"}

def analyze(path, decode_window=False):
    with open(path,"r",encoding="utf-8") as f: d=json.load(f)
    evs=d["traceEvents"]
    lo,hi=-1e30,1e30
    if decode_window:
        dc=[(e["ts"],e["ts"]+e.get("dur",0)) for e in evs
            if e.get("cat")=="gpu_user_annotation" and (e.get("name","")).startswith("decode[")]
        lo,hi=min(a for a,b in dc),max(b for a,b in dc)
    catt=collections.Counter(); knames=collections.defaultdict(lambda: collections.Counter())
    kcount=collections.defaultdict(lambda: collections.Counter()); busy=0.0
    for e in evs:
        if e.get("cat") in KCAT:
            ts=e.get("ts")
            if ts is None or ts<lo or ts>hi: continue
            dur=e.get("dur",0); nm=e.get("name","")
            c=cat_of(nm); catt[c]+=dur; knames[c][nm]+=dur; kcount[c][nm]+=1; busy+=dur
    return {"catt":catt,"knames":knames,"kcount":kcount,"busy":busy}

def cnt_sub(D,sub):
    n=0
    for c in D["kcount"]:
        for nm,k in D["kcount"][c].items():
            if sub in nm.lower(): n+=k
    return n

def steps_of(D,known=None):
    if known: return known
    g=cnt_sub(D,"topkgatingsoftmax")
    if g: return round(g/60)
    r=cnt_sub(D,"routingindicescluster")
    if r: return round(r/60)
    gd=cnt_sub(D,"gated_delta_rule_packed")
    if gd: return round(gd/45)
    return 1

def short_kernel(nm):
    n=nm
    if n.startswith("Cijk") or "Cijk_" in n: return "Tensile Cijk GEMM"
    if n.startswith("void "): n=n[5:]
    if n.startswith("std::enable_if"): return "ck_tile fmha (gfx950)"
    low=n.lower()
    table=[
        ("lamport","flashinfer twoshot allreduce+rmsNorm(Lamport)"),
        ("twoshotallreduce","flashinfer twoshot allreduce"),
        ("cross_device_reduce","aiter cross_device_reduce_1stage"),
        ("nccldevkernel_allgather","ncclDevKernel AllGather"),
        ("nccldevkernel","ncclDevKernel (NCCL)"),
        ("allgather","allgather"),
        ("bmm_e2m1_e2m1e2m1","trtllm bmm MoE gemm1 (mxfp4 experts)"),
        ("bmm_bfloat16_e2m1e2m1","trtllm bmm MoE gemm2 (mxfp4 experts)"),
        ("bmm_","trtllm bmm MoE expert gemm"),
        ("finalizekernel","cutlass moe finalize"),
        ("routingindicescluster","trtllm moe routing cluster"),
        ("routingindicesblock","trtllm moe routing block"),
        ("routing","trtllm moe routing"),
        ("moe_mxgemm","ck moe_mxgemm_2lds"),
        ("mfma_moe1","mfma_moe1_silu_mul_afp4_wfp4"),
        ("mfma_moe2","mfma_moe2_afp4_wfp4_cshuffle"),
        ("topkgatingsoftmax","vllm topkGatingSoftmax"),
        ("mxfp4_quant_moe_sort","aiter mxfp4_quant_moe_sort"),
        ("fused_mx_quant_moe_sort","aiter fused_mx_quant_moe_sort"),
        ("moesorting","ck_tile MoeSorting"),
        ("nvjet","nvjet sm100 GEMM"),
        ("splitkreduce","cublasLt splitK reduce"),
        ("hgemm_bf16","aiter hgemm_bf16 (splitK)"),
        ("fmhasm100","flashinfer fmha sm100 (paged, gen)"),
        ("paged_attention_decode_sliding_window","paged_attention_decode_sliding_window"),
        ("paged_attention_decode_ps_reduce","paged_attention ps_reduce"),
        ("unified_attention","kernel_unified_attention_3d"),
        ("reduce_segments","unified_attn reduce_segments"),
        ("fused_recurrent_gated_delta_rule_packed","fused_recurrent_gated_delta_rule_packed_decode"),
        ("fused_recurrent_gated_delta_rule","fused_recurrent_gated_delta_rule_fwd"),
        ("causal_conv1d_update","causal_conv1d_update"),
        ("act_and_mul","act_and_mul (silu)"),
        ("layer_norm_fwd","layer_norm_fwd"),
        ("rmsnormkernel","flashinfer rmsNorm"),
        ("mean_mul_pow_rsqrt_silu","triton rmsnorm+silu"),
        ("mean_mul_pow_rsqrt","triton rmsnorm"),
        ("fused_gate_sigmoid_mul_add","fused_gate_sigmoid_mul_add"),
        ("fused_sigmoid_mul","fused_sigmoid_mul"),
        ("all_reduce__mul_sigmoid","triton post-allreduce residual"),
        ("qkvzba_split_reshape_cat","fused_qkvzba_split_reshape_cat"),
        ("nvfp4_quantize","flashinfer nvfp4 quantize"),
        ("per_tensor_quant_fp8","per_tensor_quant_fp8"),
        ("per_tensor_absmax","per_tensor_absmax"),
        ("fp8_set_kv_buffer","fused_fp8_set_kv_buffer"),
        ("reshape_and_cache","reshape_and_cache + quant"),
        ("memcpy128","Memcpy128"),
        ("copybuffer","rocclr copyBuffer"),
        ("memcpy dtod","Memcpy DtoD"),
        ("mix_sample","aiter mix_sample"),
        ("greedy_sample","greedy_sample"),
        ("argmax","argmax sample"),
        ("mrope","triton mrope"),
        ("direct_copy","elementwise copy"),
    ]
    for key,label in table:
        if key in low: return label
    for ch in ("<","("):
        if ch in n: n=n.split(ch)[0]
    return n[:38]

def kernel_list(D,cat,topn=5):
    if cat not in D["knames"]: return ""
    seen=[]
    for nm,t in D["knames"][cat].most_common():
        s=short_kernel(nm).strip()
        if s and s not in seen: seen.append(s)
        if len(seen)>=topn: break
    return "\n".join(f"• {s}" for s in seen)

print("loading Atom...",flush=True);  A=analyze(ATOM,decode_window=True)
print("loading MI355...",flush=True); M=analyze(SGL)
print("loading B200...",flush=True);  B=analyze(B200)
A_ST=steps_of(A,known=1024); M_ST=steps_of(M); B_ST=steps_of(B)
print(f"steps: Atom={A_ST} MI355={M_ST} B200={B_ST}")

wb=openpyxl.Workbook()
HDR=PatternFill("solid",fgColor="305496"); HDRF=Font(bold=True,color="FFFFFF",size=11)
SUB=PatternFill("solid",fgColor="D9E1F2"); TITLE=Font(bold=True,size=14,color="1F4E78")
GOOD=PatternFill("solid",fgColor="C6EFCE"); BAD=PatternFill("solid",fgColor="FFC7CE")
BOLD=Font(bold=True); thin=Side(style="thin",color="BFBFBF"); BORDER=Border(thin,thin,thin,thin)
CENTER=Alignment(horizontal="center"); LEFT=Alignment(horizontal="left",vertical="center")
WRAP=Alignment(horizontal="left",vertical="top",wrap_text=True)
def hdr(ws,row,cols,widths):
    for i,c in enumerate(cols,1):
        cell=ws.cell(row=row,column=i,value=c); cell.fill=HDR; cell.font=HDRF; cell.alignment=CENTER; cell.border=BORDER
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

order=["moe","gemm","mixer:linear(gdn)","mixer:full-attn","communication","normalization","activation","quantization","elementwise","memcpy","other"]

# Sheet 1: single-step 3-way gap
ws=wb.active; ws.title="SingleStep_Gap"
ws["A1"]="Qwen3.5-397B-A17B Decode — per single step: sglang MI355 vs sglang B200 (Atom=AMD ref), bs=128"; ws["A1"].font=TITLE
ws["A2"]=(f"µs/step = category busy / steps (Atom {A_ST}, MI355 {M_ST}, B200 {B_ST}). "
          f"Gap = MI355 / B200 (>1 => MI355 slower). Naive-sum busy (multi-stream overlap not subtracted).")
ws["A2"].font=Font(italic=True,size=9,color="808080")
r=4
hdr(ws,r,["Category","Atom µs/step","MI355 µs/step","B200 µs/step","Gap MI355/B200","B200 kernels used","MI355 kernels used"],
    [20,13,14,13,15,44,44])
r+=1
for c in order:
    a=A["catt"].get(c,0)/A_ST; m=M["catt"].get(c,0)/M_ST; b=B["catt"].get(c,0)/B_ST
    gap=(m/b) if b>0 else 0
    ws.cell(row=r,column=1,value=c).font=BOLD; ws.cell(row=r,column=1).alignment=LEFT
    for i,v in enumerate([round(a,1),round(m,1),round(b,1)],2):
        cell=ws.cell(row=r,column=i,value=v); cell.alignment=CENTER; cell.number_format='#,##0.0'
    gc=ws.cell(row=r,column=5,value=round(gap,2) if gap else "n/a"); gc.alignment=CENTER; gc.number_format='0.00"x"'
    if gap>1.15: gc.fill=BAD
    elif 0<gap<0.87: gc.fill=GOOD
    ws.cell(row=r,column=6,value=kernel_list(B,c)).alignment=WRAP
    ws.cell(row=r,column=7,value=kernel_list(M,c)).alignment=WRAP
    for i in range(1,8): ws.cell(row=r,column=i).border=BORDER
    r+=1
ta=sum(A["catt"].values())/A_ST; tm=sum(M["catt"].values())/M_ST; tb=sum(B["catt"].values())/B_ST
ws.cell(row=r,column=1,value="TOTAL µs/step").font=BOLD; ws.cell(row=r,column=1).fill=SUB
for i,v in enumerate([round(ta,1),round(tm,1),round(tb,1),round(tm/tb,2)],2):
    cell=ws.cell(row=r,column=i,value=v); cell.fill=SUB; cell.alignment=CENTER
    cell.number_format='0.00"x"' if i==5 else '#,##0.0'
for i in range(1,8):
    ws.cell(row=r,column=i).border=BORDER
    if ws.cell(row=r,column=i).value is None: ws.cell(row=r,column=i).fill=SUB
r+=2

# Gap summary notes
ws.cell(row=r,column=1,value="Gap read-out").font=TITLE; r+=1
notes=[
 f"TOTAL: MI355 {tm/1000:.1f} ms/step vs B200 {tb/1000:.1f} ms/step  ->  MI355 is {tm/tb:.2f}x B200.",
 "MoE (~45-49% of step) is the dominant term on both; B200 uses trtllm bmm(mxfp4) + cutlass routing/finalize, MI355 uses ck/mfma moe. This is the #1 gap driver.",
 "GEMM: B200 nvjet sm100 + cublas splitK vs MI355 Tensile Cijk. Compare per-step to see dense-gemm efficiency.",
 "communication: B200 flashinfer twoshot(Lamport) allreduce vs MI355 aiter cross_device_reduce_1stage.",
 "full-attn: B200 flashinfer fmha sm100 vs MI355 kernel_unified_attention_3d; gdn: both fused_recurrent_gated_delta packed decode.",
 "Where Gap>1.15 (red) MI355 is slower and is the highest-ROI target; Gap<0.87 (green) MI355 already wins.",
]
for t in notes:
    ws.cell(row=r,column=1,value=("• "+t)); ws.cell(row=r,column=1).alignment=WRAP; r+=1

# Sheet 2: kernel dictionary per engine
ws2=wb.create_sheet("Kernel_Map")
ws2["A1"]="Per-category kernels actually used (top by time)"; ws2["A1"].font=TITLE
r=3; hdr(ws2,r,["Category","B200 (NVIDIA)","MI355 (AMD/sglang)","Atom (AMD)"],[20,46,46,46]); r+=1
for c in order:
    ws2.cell(row=r,column=1,value=c).font=BOLD; ws2.cell(row=r,column=1).alignment=LEFT
    ws2.cell(row=r,column=2,value=kernel_list(B,c,6)).alignment=WRAP
    ws2.cell(row=r,column=3,value=kernel_list(M,c,6)).alignment=WRAP
    ws2.cell(row=r,column=4,value=kernel_list(A,c,6)).alignment=WRAP
    for i in range(1,5): ws2.cell(row=r,column=i).border=BORDER
    r+=1

wb.save(OUT); print("saved:",OUT)
