import json, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ATOM = r"C:\Workspace\qwen\Atom_Qwen3.5-397B-A17B-MXFP4_ts_20260728_101711_991.pt.trace.json"
SGL  = r"C:\Workspace\qwen\profile_qwen_mi355_8k1k_c128_1784806436.739748\qwen_mi355_8k1k_c128-1784806436.7420862-TP-0-DECODE.trace.json"
OUT  = r"C:\Workspace\qwen\atom_vs_sglang_decode.xlsx"

def cat_of(name):
    n = name.lower()
    is_tri = n.startswith("triton") or "fused" in n
    if ("cross_device_reduce" in n or "nccl" in n or "rccl" in n or "all_gather" in n or "allgather" in n
        or "reduce_scatter" in n or "quickreduce" in n or "custom_ar" in n
        or (("all_reduce" in n or "allreduce" in n) and not is_tri)):
        return "communication"
    if "moe" in n or "topkgating" in n or "expert" in n or "grouped_gemm" in n or "group_gemm" in n:
        return "moe"
    if any(k in n for k in ["gated_delta","gdn","causal_conv","recurrent_gated","chunk_gated","split_chunk","chunk_fwd","recompute_w_u"]):
        return "mixer:linear(gdn)"
    if any(k in n for k in ["fmha","flash_attn","_attn_","paged_attn","paged_attention","mha","attention","_attn"]):
        return "mixer:full-attn"
    if n.startswith("cijk") or "gemm" in n or "_mm_" in n or "matmul" in n or "hgemm" in n or "nvjet" in n or "cutlass" in n:
        return "gemm"
    if "rsqrt" in n or "rmsnorm" in n or "layernorm" in n or "qk_norm" in n or "layer_norm" in n:
        return "normalization"
    if "quant" in n or "fp4" in n or "fp8" in n or "preshuffle" in n or "scaled" in n:
        return "quantization"
    if any(k in n for k in ["act_and_mul","silu","gelu","sigmoid","softmax"]):
        return "activation"
    if any(k in n for k in ["copybuffer","memcpy","memset","__amd_rocclr","dtod","htod"]):
        return "memcpy"
    if any(k in n for k in ["elementwise","embedding","rope","rotary","index","cat","slice","add","mul","fill","_to_copy"]):
        return "elementwise"
    return "other"

KCAT = {"kernel","gpu_memcpy","gpu_memset"}

def analyze(path, decode_window=False):
    with open(path,"r",encoding="utf-8") as f:
        d=json.load(f)
    evs=d["traceEvents"]
    lo,hi=-1e30,1e30
    if decode_window:
        dc=[(e["ts"],e["ts"]+e.get("dur",0)) for e in evs
            if e.get("cat")=="gpu_user_annotation" and (e.get("name","")).startswith("decode[")]
        lo,hi=min(a for a,b in dc),max(b for a,b in dc)
    catt=collections.Counter(); catc=collections.Counter()
    knames=collections.defaultdict(lambda: collections.Counter())
    kcount=collections.defaultdict(lambda: collections.Counter())
    busy=0.0
    for e in evs:
        if e.get("cat") in KCAT:
            ts=e.get("ts")
            if ts is None or ts<lo or ts>hi: continue
            dur=e.get("dur",0); nm=e.get("name","")
            c=cat_of(nm); catt[c]+=dur; catc[c]+=1; knames[c][nm]+=dur; kcount[c][nm]+=1; busy+=dur
    return {"catt":catt,"catc":catc,"knames":knames,"kcount":kcount,"busy":busy}

print("loading Atom (decode window)...",flush=True)
A=analyze(ATOM,decode_window=True)
print("loading sglang MI355 c128 decode...",flush=True)
S=analyze(SGL,decode_window=False)

# ---------------- decode step counts ----------------
N_LAYERS = 60  # all 60 decoder layers run MoE gating once per step
def gating_count(D):
    n=0
    for c in D["kcount"]:
        for nm,cnt in D["kcount"][c].items():
            if "topkgatingsoftmax" in nm.lower():
                n+=cnt
    return n
A_gate=gating_count(A); S_gate=gating_count(S)
ATOM_STEPS=1024  # from GPU decode[...] annotations
SGL_STEPS=max(1, round(S_gate/N_LAYERS))
print(f"Atom  topkGating count={A_gate} -> implied steps={A_gate/N_LAYERS:.1f} (using known {ATOM_STEPS})")
print(f"sglang topkGating count={S_gate} -> steps={SGL_STEPS}")
A_STEPS=ATOM_STEPS

def short_kernel(nm):
    n=nm
    if n.startswith("Cijk") or "Cijk_" in n:
        return "Tensile Cijk GEMM"
    for pre in ("void ",):
        if n.startswith(pre): n=n[len(pre):]
    # strip std::enable_if wrapper
    if n.startswith("std::enable_if"):
        return "ck_tile fmha (gfx950)"
    # demangle-ish for _ZN... mangled names: pull readable token
    low=n.lower()
    table=[
        ("cross_device_reduce","aiter cross_device_reduce_1stage"),
        ("allgather","aiter allgather"),
        ("nccldevkernel","ncclDevKernel (RCCL)"),
        ("moe_mxgemm","ck moe_mxgemm_2lds"),
        ("mfma_moe1","mfma_moe1_silu_mul_afp4_wfp4"),
        ("mfma_moe2","mfma_moe2_afp4_wfp4_cshuffle"),
        ("topkgatingsoftmax","vllm topkGatingSoftmax"),
        ("mxfp4_quant_moe_sort","aiter mxfp4_quant_moe_sort"),
        ("fused_mx_quant_moe_sort","aiter fused_mx_quant_moe_sort"),
        ("moesorting","ck_tile MoeSorting"),
        ("opus_moe_sorting","aiter opus_moe_sorting"),
        ("paged_attention_decode_sliding_window","paged_attention_decode_sliding_window"),
        ("paged_attention_decode_ps_reduce","paged_attention ps_reduce"),
        ("unified_attention","kernel_unified_attention_3d"),
        ("reduce_segments","unified_attn reduce_segments"),
        ("fused_recurrent_gated_delta_rule_packed","fused_recurrent_gated_delta_rule_packed_decode"),
        ("fused_recurrent_gated_delta_rule","fused_recurrent_gated_delta_rule_fwd"),
        ("causal_conv1d_update","causal_conv1d_update"),
        ("fused_split_chunk","fused_split_chunk"),
        ("fused_gdn_gating","fused_gdn_gating"),
        ("act_and_mul","act_and_mul (silu)"),
        ("gemma_fused_add_rmsnorm","fused_add_rmsnorm"),
        ("layer_norm_fwd","layer_norm_fwd"),
        ("fused_qk_norm","fused_qk_norm"),
        ("mean_mul_pow_rsqrt_silu","triton rmsnorm+silu"),
        ("mean_mul_pow_rsqrt","triton rmsnorm"),
        ("fused_gate_sigmoid_mul_add","fused_gate_sigmoid_mul_add"),
        ("fused_sigmoid_mul","fused_sigmoid_mul"),
        ("all_reduce__mul_sigmoid","triton post-allreduce residual"),
        ("qkvzba_split_reshape_cat","fused_qkvzba_split_reshape_cat"),
        ("split_squeeze","triton cat/slice glue"),
        ("cat_slice","triton cat/slice glue"),
        ("reshape_and_cache","reshape_and_cache + quant"),
        ("copybuffer","rocclr copyBuffer"),
        ("memcpy dtod","Memcpy DtoD"),
        ("mix_sample","aiter mix_sample"),
        ("greedy_sample","aiter greedy_sample"),
        ("hgemm_bf16","aiter hgemm_bf16 (splitK)"),
        ("mrope","triton mrope"),
        ("wv_splitk","aiter wv_splitk"),
    ]
    for key,label in table:
        if key in low:
            return label
    # generic: cut at first '<' or '('
    for ch in ("<","("):
        if ch in n: n=n.split(ch)[0]
    return n[:40]

def kernel_list(D, cat, topn=5):
    if cat not in D["knames"]: return ""
    seen=[]; 
    for nm,t in D["knames"][cat].most_common():
        s=short_kernel(nm).strip()
        if s and s not in seen:
            seen.append(s)
        if len(seen)>=topn: break
    return "\n".join(f"• {s}" for s in seen)

# ---------------- Excel styling ----------------
wb=openpyxl.Workbook()
HDR=PatternFill("solid",fgColor="305496"); HDRF=Font(bold=True,color="FFFFFF",size=11)
SUB=PatternFill("solid",fgColor="D9E1F2"); TITLE=Font(bold=True,size=14,color="1F4E78")
CATFILL=PatternFill("solid",fgColor="FCE4D6"); BOLD=Font(bold=True)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(thin,thin,thin,thin)
CENTER=Alignment(horizontal="center"); LEFT=Alignment(horizontal="left",vertical="center")

def hdr_row(ws,row,cols,widths=None):
    for i,c in enumerate(cols,1):
        cell=ws.cell(row=row,column=i,value=c); cell.fill=HDR; cell.font=HDRF
        cell.alignment=CENTER; cell.border=BORDER
    if widths:
        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w

# ============ Sheet 1: Category_Comparison (per single decode step) ============
ws=wb.active; ws.title="Category_Comparison"
ws["A1"]="Qwen3.5-397B-A17B  Decode  —  Atom vs sglang, per single step (MI355-class AMD, bs=128)"; ws["A1"].font=TITLE
ws["A2"]=(f"Per-step = category GPU busy / decode steps  (Atom {A_STEPS} steps; sglang {SGL_STEPS} steps, "
          f"derived from topkGatingSoftmax count / {N_LAYERS} layers). Last two columns list the kernels actually used.")
ws["A2"].font=Font(italic=True,size=9,color="808080")
r=4
hdr_row(ws,r,["Category","Atom µs/step","sglang µs/step","Atom %","sglang %","Atom kernels used","sglang kernels used"],
        [20,14,15,10,10,42,42])
order=["moe","gemm","mixer:linear(gdn)","mixer:full-attn","communication","normalization","activation","quantization","elementwise","memcpy","other"]
WRAP=Alignment(horizontal="left",vertical="top",wrap_text=True)
r+=1
for c in order:
    a_us=A["catt"].get(c,0)/A_STEPS; s_us=S["catt"].get(c,0)/SGL_STEPS
    ap=A["catt"].get(c,0)/A["busy"]*100; sp=S["catt"].get(c,0)/S["busy"]*100
    ws.cell(row=r,column=1,value=c).font=BOLD; ws.cell(row=r,column=1).alignment=LEFT
    vals=[round(a_us,1),round(s_us,1),round(ap,1),round(sp,1)]
    for i,v in enumerate(vals,2):
        cell=ws.cell(row=r,column=i,value=v); cell.alignment=CENTER
        if i in (4,5): cell.number_format='0.0"%"'
        else: cell.number_format='#,##0.0'
    ws.cell(row=r,column=6,value=kernel_list(A,c)).alignment=WRAP
    ws.cell(row=r,column=7,value=kernel_list(S,c)).alignment=WRAP
    for i in range(1,8): ws.cell(row=r,column=i).border=BORDER
    r+=1
# mixer total + per-step total row
am_us=(A["catt"].get('mixer:linear(gdn)',0)+A["catt"].get('mixer:full-attn',0))/A_STEPS
sm_us=(S["catt"].get('mixer:linear(gdn)',0)+S["catt"].get('mixer:full-attn',0))/SGL_STEPS
tot_a=sum(A["catt"].values())/A_STEPS; tot_s=sum(S["catt"].values())/SGL_STEPS
for label,vals in [("token-mixer TOTAL",[round(am_us,1),round(sm_us,1),"",""]),
                   ("TOTAL µs/step",[round(tot_a,1),round(tot_s,1),"",""])]:
    ws.cell(row=r,column=1,value=label).font=BOLD; ws.cell(row=r,column=1).fill=SUB
    for i,v in enumerate(vals,2):
        cell=ws.cell(row=r,column=i,value=v); cell.fill=SUB; cell.alignment=CENTER
        if isinstance(v,(int,float)): cell.number_format='#,##0.0'
    for i in range(1,8):
        ws.cell(row=r,column=i).border=BORDER
        if ws.cell(row=r,column=i).fill.fgColor.rgb!="00D9E1F2": ws.cell(row=r,column=i).fill=SUB
    r+=1

# ============ Sheet 2: Layer_Decomposition ============
ws2=wb.create_sheet("Layer_Decomposition")
ws2["A1"]="Decode decomposition by model layer structure"; ws2["A1"].font=TITLE
ws2["A2"]="Model = 60 decoder layers: 45x LinearDecoderLayer (GatedDeltaNet mixer) + 15x AttentionDecoderLayer (full-attn mixer). Both share MoE + GEMM + norm + TP-comm."
ws2["A2"].font=Font(italic=True,size=9,color="808080")
r=4
hdr_row(ws2,r,["Component","Scope","Atom %","sglang %"],[26,26,12,12])
r+=1
rows=[
 ("channel-mixer = MoE","shared by both layer types","moe"),
 ("projection GEMM (qkv/o/gate/up/down)","shared by both layer types","gemm"),
 ("TP communication (allreduce/allgather)","shared (2x per layer)","communication"),
 ("normalization (RMSNorm/qk_norm)","shared","normalization"),
 ("activation (silu/gate)","shared","activation"),
 ("token-mixer: GatedDeltaNet (linear)","LinearDecoderLayer only (x45)","mixer:linear(gdn)"),
 ("token-mixer: full attention","AttentionDecoderLayer only (x15)","mixer:full-attn"),
]
for name,scope,c in rows:
    ap=A["catt"].get(c,0)/A["busy"]*100; sp=S["catt"].get(c,0)/S["busy"]*100
    ws2.cell(row=r,column=1,value=name).alignment=LEFT
    ws2.cell(row=r,column=2,value=scope).alignment=LEFT
    ca=ws2.cell(row=r,column=3,value=round(ap,1)); cs=ws2.cell(row=r,column=4,value=round(sp,1))
    ca.number_format='0.0"%"'; cs.number_format='0.0"%"'
    for i in range(1,5):
        ws2.cell(row=r,column=i).border=BORDER
        if i>2: ws2.cell(row=r,column=i).alignment=CENTER
    if "token-mixer" in name:
        for i in range(1,5): ws2.cell(row=r,column=i).fill=CATFILL
    r+=1

# ============ Sheets 3&4: kernel reference lists ============
def kernel_sheet(name, D, title):
    ws=wb.create_sheet(name)
    ws["A1"]=title; ws["A1"].font=TITLE
    r=3
    hdr_row(ws,r,["Category","Kernel name","Time (ms)","Share %","Count"],[20,88,12,10,12])
    r+=1
    for c in order:
        if c not in D["knames"]: continue
        cat_total=sum(D["knames"][c].values())
        # category header row
        cell=ws.cell(row=r,column=1,value=c); cell.font=BOLD; cell.fill=CATFILL; cell.border=BORDER
        cell2=ws.cell(row=r,column=3,value=round(cat_total/1000,1)); cell2.font=BOLD; cell2.fill=CATFILL; cell2.number_format='#,##0.0'; cell2.border=BORDER
        cell3=ws.cell(row=r,column=4,value=round(cat_total/D["busy"]*100,1)); cell3.font=BOLD; cell3.fill=CATFILL; cell3.number_format='0.0"%"'; cell3.border=BORDER
        ws.cell(row=r,column=2).fill=CATFILL; ws.cell(row=r,column=2).border=BORDER
        ws.cell(row=r,column=5).fill=CATFILL; ws.cell(row=r,column=5).border=BORDER
        r+=1
        for nm,t in D["knames"][c].most_common(12):
            ws.cell(row=r,column=1,value="")
            ws.cell(row=r,column=2,value=nm).alignment=LEFT
            ct=ws.cell(row=r,column=3,value=round(t/1000,1)); ct.number_format='#,##0.0'; ct.alignment=CENTER
            cp=ws.cell(row=r,column=4,value=round(t/D["busy"]*100,2)); cp.number_format='0.00"%"'; cp.alignment=CENTER
            cc=ws.cell(row=r,column=5,value=D["kcount"][c][nm]); cc.number_format='#,##0'; cc.alignment=CENTER
            for i in range(1,6): ws.cell(row=r,column=i).border=BORDER
            r+=1
    ws.freeze_panes="A4"

kernel_sheet("Atom_Kernels", A, "Atom decode — actual kernels used (top 12 per category)")
kernel_sheet("Sglang_Kernels", S, "sglang MI355 c128 decode — actual kernels used (top 12 per category)")

for ws_ in wb.worksheets:
    ws_.sheet_view.showGridLines=False
wb.save(OUT)
print("saved:",OUT)
