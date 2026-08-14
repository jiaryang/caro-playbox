#!/usr/bin/env python
"""Build a consolidated Excel report comparing MI355 vs B200 across
c4/c128 x prefill/decode. Reuses trace_toolbox.analyze().

Sheets:
  Summary   - e2e wall / kernel-sum / gpu-active / concurrency for all 8 runs
  Categories- category ms/step rollup, B200 vs MI355 + ratio, per config
  <run>     - one per (platform x config): every kernel in execution order
"""
import os, sys, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import trace_toolbox as tt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

QDIR = r"C:\Workspace\qwen"

# (config, phase, platform) -> (relative path, analyze-phase)
D_C4  = "profile_qwen_b200_8k1k_c4_1784803611.7704022"
M_C4  = "profile_qwen_mi355_8k1k_c4_1784806235.0348327"
D_128 = "profile_qwen_b200_8k1k_c128_1784800637.471206"
M_128 = "profile_qwen_mi355_8k1k_c128_1784806436.739748"

RUNS = [
    # config, phase, platform, path, analyze_phase
    ("c4",   "decode",  "B200",  f"{D_C4}\\qwen_b200_8k1k_c4-1784803611.7740567-TP-0-DECODE.trace.json",  "decode"),
    ("c4",   "decode",  "MI355", f"{M_C4}\\qwen_mi355_8k1k_c4-1784806235.0370286-TP-0-DECODE.trace.json", "decode"),
    ("c4",   "prefill", "B200",  f"{D_C4}\\qwen_b200_8k1k_c4-1784803611.7740567-TP-0-EXTEND.trace.json",  "prefill"),
    ("c4",   "prefill", "MI355", f"{M_C4}\\qwen_mi355_8k1k_c4-1784806235.0370286-TP-0-EXTEND.trace.json", "prefill"),
    ("c128", "decode",  "B200",  f"{D_128}\\qwen_b200_8k1k_c128-1784800637.475314-TP-0-DECODE.trace.json", "decode"),
    ("c128", "decode",  "MI355", f"{M_128}\\qwen_mi355_8k1k_c128-1784806436.7420862-TP-0-DECODE.trace.json","decode"),
    ("c128", "prefill", "B200",  f"{D_128}\\qwen_b200_8k1k_c128-1784800637.475314-TP-0-EXTEND.trace.json", "prefill"),
    ("c128", "prefill", "MI355", f"{M_128}\\qwen_mi355_8k1k_c128-1784806436.7420862-TP-0-EXTEND.trace.json","prefill"),
]

# (config, platform) -> serving-benchmark log file (client-side latency metrics)
LOGS = {
    ("c4",   "B200"):  "B200_qwen_8192_o1024_c4.log",
    ("c4",   "MI355"): "MI355_qwen_8192_o1024_c4.log",
    ("c128", "B200"):  "B200_qwen_8192_o1024_c128.log",
    ("c128", "MI355"): "MI355_qwen_8192_o1024_c128.log",
}

def parse_log(path):
    """Extract client-side latency metrics (ms) from a serving-benchmark log."""
    want = {
        "e2e_mean":    r"Mean E2E Latency \(ms\):",
        "e2e_med":     r"Median E2E Latency \(ms\):",
        "ttft_mean":   r"Mean TTFT \(ms\):",
        "ttft_med":    r"Median TTFT \(ms\):",
        "tpot_mean":   r"Mean TPOT \(ms\):",
        "tpot_med":    r"Median TPOT \(ms\):",
        "itl_mean":    r"Mean ITL \(ms\):",
        "itl_med":     r"Median ITL \(ms\):",
    }
    out = {}
    if not os.path.exists(path):
        return out
    with open(path, encoding="utf-8", errors="ignore") as f:
        text = f.read()
    for k, pat in want.items():
        m = re.search(pat + r"\s*([0-9.]+)", text)
        if m:
            out[k] = float(m.group(1))
    return out

# (config, phase, platform) -> analyzer detail workbook (layer/module level).
# Only c4 was exported by trace_module_analyzer.py.
DETAIL = {
    ("c4",   "decode",  "B200"):  "detail_b200_c4_decode.xlsx",
    ("c4",   "decode",  "MI355"): "detail_mi355_c4_decode.xlsx",
    ("c4",   "prefill", "B200"):  "detail_b200_c4_prefill.xlsx",
    ("c4",   "prefill", "MI355"): "detail_mi355_c4_prefill.xlsx",
    ("c128", "decode",  "B200"):  "detail_b200_c128_decode.xlsx",
    ("c128", "decode",  "MI355"): "detail_mi355_c128_decode.xlsx",
    ("c128", "prefill", "B200"):  "detail_b200_c128_prefill.xlsx",
    ("c128", "prefill", "MI355"): "detail_mi355_c128_prefill.xlsx",
}

_TREE = str.maketrans("", "", "└├─│")

def _pct(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).rstrip("% "))
    except ValueError:
        return None

def read_detail(path):
    """Read a trace_module_analyzer Summary sheet -> total_us, layer rollup, categories."""
    out = {"total_us": None, "layers": [], "cats": []}
    if not os.path.exists(path):
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Summary"].iter_rows(values_only=True))
    mt_rows = list(wb["Module Tree"].iter_rows(values_only=True)) if "Module Tree" in wb.sheetnames else []
    wb.close()
    lay_re = re.compile(r"^\(([A-Za-z]+)\)\s*(\d+)x")     # decode: "(A) 44x"
    in_cat = False
    for r in rows:
        c0 = r[0]
        if c0 is None:
            in_cat = False; continue
        s0 = str(c0).translate(_TREE).strip()
        if s0 == "Total Kernel Time" and r[1] is not None:
            out["total_us"] = float(r[1]); continue
        if s0 == "Category":
            in_cat = True; continue
        if in_cat:
            # cat | count | total_us | avg | pct
            if len(r) > 2 and r[2] is not None:
                out["cats"].append((s0, float(r[2]), _pct(r[4] if len(r) > 4 else None)))
            continue
        m = lay_re.match(s0)
        is_mod = s0.startswith("Qwen3_5") and s0.endswith("Layer")
        if (m or is_mod) and r[1] is not None:
            label = f"({m.group(1)})" if m else s0
            out["layers"].append((label, _pct(r[2])))

    # Fallback: some traces (e.g. c128 prefill) roll up the Summary only to the
    # top ForCausalLM module. Recover the decoder-layer split from Module Tree
    # (which fully expands one representative forward); use share within it.
    if not out["layers"] and mt_rows:
        typ_re = re.compile(r"(Qwen3_5\w*DecoderLayer)_\d+$")
        agg = {}
        for r in mt_rows:
            if not r or r[0] is None:
                continue
            s0 = str(r[0]).translate(_TREE).strip()
            m = typ_re.match(s0)
            if m and len(r) > 1 and isinstance(r[1], (int, float)):
                agg[m.group(1)] = agg.get(m.group(1), 0.0) + float(r[1])
        fwd_tot = sum(agg.values())
        for t, us in sorted(agg.items(), key=lambda kv: -kv[1]):
            out["layers"].append((t, (us/fwd_tot*100 if fwd_tot else None)))
    return out

HF   = PatternFill("solid", fgColor="305496")
HF2  = PatternFill("solid", fgColor="8EAADB")
HFONT= Font(color="FFFFFF", bold=True)
NUM  = '0.00'      # values: 2 decimals
PCT  = '0.0"%"'    # ratios/percentages: 1 decimal
RATIO= '0.0'       # bare ratios (x): 1 decimal

def style_header(ws, row=1):
    for c in ws[row]:
        c.fill = HF; c.font = HFONT; c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row+1, 1).coordinate

def main():
    results = {}  # (config,phase,platform) -> dict(agg, rows, cat)
    for config, phase, plat, rel, aphase in RUNS:
        path = os.path.join(QDIR, rel)
        print(f"parsing {plat:5s} {config:4s} {phase:7s} ...", flush=True)
        _w, _g, _o, sections, _i = tt.analyze(path, phase=aphase)
        if not sections:
            print("   !! no windows"); continue
        lbl, agg, rows = sections[0]
        cat = {}
        for r in rows:
            cat[r["category"]] = cat.get(r["category"], 0.0) + r["tps"]
        results[(config, phase, plat)] = dict(lbl=lbl, agg=agg, rows=rows, cat=cat)

    bench = {}  # (config, platform) -> metrics dict
    for (config, plat), fname in LOGS.items():
        bench[(config, plat)] = parse_log(os.path.join(QDIR, fname))

    detail = {}  # (config, phase, platform) -> analyzer rollup
    for key, fname in DETAIL.items():
        detail[key] = read_detail(os.path.join(QDIR, fname))

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ---------------- Summary ----------------
    ws = wb.create_sheet("Summary")
    ws.append(["Qwen 8k/1k  -  MI355 vs B200  (TP-0, per-step averages)"])
    ws.append(["ms/step; E2E = wall-clock; Kernel-sum = sum of kernel durations; "
               "GPU-active = union (busy) time; Concurrency = Kernel-sum / GPU-active"])
    ws.append([])
    hdr = ["Config", "Phase", "Platform", "Steps", "E2E wall (ms)",
           "Kernel-sum (ms)", "GPU-active (ms)", "Concurrency x", "Kernels/step",
           "E2E vs B200 (%)"]
    ws.append(hdr); hrow = ws.max_row
    for config in ("c4", "c128"):
        for phase in ("prefill", "decode"):
            b = results.get((config, phase, "B200"))
            base_wall = b["agg"]["wall"] if b else None
            for plat in ("B200", "MI355"):
                r = results.get((config, phase, plat))
                if not r: continue
                a = r["agg"]; conc = a["work"]/a["union"] if a["union"] else 0
                ratio = (a["wall"]/base_wall*100) if base_wall else None
                ws.append([config, phase, plat, a["n"],
                           a["wall"]/1000, a["work"]/1000,
                           a["union"]/1000, conc,
                           round(sum(a["cnt"].values())/a["n"], 0) if a["n"] else 0,
                           ratio])
                for col in (5, 6, 7):
                    ws.cell(ws.max_row, col).number_format = NUM
                ws.cell(ws.max_row, 8).number_format = RATIO
                ws.cell(ws.max_row, 10).number_format = PCT
                if plat == "MI355":
                    ws.cell(ws.max_row, 10).font = Font(bold=True,
                        color="C00000" if (ratio and ratio > 100) else "008000")
    ws.cell(1, 1).font = Font(bold=True, size=13)
    for c in ws[hrow]:
        c.fill = HF; c.font = HFONT; c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(hrow+1, 1).coordinate
    for i, w in enumerate([9, 9, 9, 8, 14, 16, 16, 14, 13, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- Bench vs Trace ----------------
    ws = wb.create_sheet("Bench_vs_Trace")
    ws.append(["Trace kernel time (3 conventions)  vs  serving-benchmark client latency"])
    ws.append(["Kernel-sum = naive sum of kernel durations (double-counts multi-stream overlap);  "
               "GPU-active (union) = de-overlapped GPU-busy time = apples-to-apples with bench;  "
               "E2E wall = step-window wall (incl. profiling idle bubbles)."])
    ws.append(["Decode: compare GPU-active(union) <-> bench ITL median (per-token GPU work).  "
               "Prefill: bench TTFT = full request incl. queue/tokenize & chunked over several forwards, "
               "so trace per-step is NOT directly comparable (trend only)."])
    ws.append([])
    hdr = ["Config", "Phase", "Platform", "Steps",
           "Kernel-sum (ms)", "GPU-active union (ms)", "E2E wall (ms)",
           "Bench TTFT mean", "Bench TTFT med", "Bench TPOT mean", "Bench TPOT med",
           "Bench ITL mean", "Bench ITL med",
           "union vs ITL-med (%)", "Aligned metric"]
    ws.append(hdr); hrow = ws.max_row
    for config in ("c4", "c128"):
        for phase in ("decode", "prefill"):
            for plat in ("B200", "MI355"):
                r = results.get((config, phase, plat))
                if not r: continue
                a = r["agg"]; bm = bench.get((config, plat), {})
                union_ms = a["union"]/1000
                itl_med = bm.get("itl_med")
                if phase == "decode":
                    align_pct = (union_ms/itl_med*100) if itl_med else None
                    align_note = "GPU-active(union) <-> ITL median"
                else:
                    align_pct = None
                    align_note = "TTFT=full req (queue+chunked); trend only"
                ws.append([config, phase, plat, a["n"],
                           a["work"]/1000, union_ms, a["wall"]/1000,
                           bm.get("ttft_mean"), bm.get("ttft_med"),
                           bm.get("tpot_mean"), bm.get("tpot_med"),
                           bm.get("itl_mean"), bm.get("itl_med"),
                           align_pct, align_note])
                for col in range(5, 14):
                    ws.cell(ws.max_row, col).number_format = NUM
                ws.cell(ws.max_row, 14).number_format = PCT
                if phase == "decode" and align_pct is not None:
                    # green when trace GPU-busy is within +/-25% of client ITL
                    ok = 75 <= align_pct <= 125
                    ws.cell(ws.max_row, 14).font = Font(bold=True,
                        color="008000" if ok else "C00000")
    ws.cell(1, 1).font = Font(bold=True, size=13)
    for i in (2, 3):
        ws.cell(i, 1).font = Font(italic=True, size=9, color="595959")
    for c in ws[hrow]:
        c.fill = HF; c.font = HFONT; c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(hrow+1, 1).coordinate
    for i, w in enumerate([9, 9, 9, 7, 15, 20, 13, 15, 14, 15, 14, 13, 12, 18, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- Drilldown (E2E -> step -> layer -> category) ----------------
    ws = wb.create_sheet("Drilldown")
    IN, OUT = 8192, 1024
    ws.append(["Drill-down:  ① E2E (bench client)  ->  ② step (trace toolbox)  ->  "
               "③ layer type (analyzer)  ->  ④ kernel category"])
    ws.append(["Same GPU time, four zoom levels. Analyzer totals span all captured forwards; "
               "normalised to per-step by dividing by detected replay count "
               "rep = round(analyzer_total / toolbox_kernel-sum_per_step)."])
    ws.append([])
    ws.append(["Level", "Item", "ms", "share %", "note"]); hrow = ws.max_row
    L1 = PatternFill("solid", fgColor="1F4E78")   # E2E
    L2 = PatternFill("solid", fgColor="2E75B6")   # step
    L3 = PatternFill("solid", fgColor="5B9BD5")   # layer
    L4 = PatternFill("solid", fgColor="9DC3E6")   # category
    WHT = Font(color="FFFFFF", bold=True)

    def drow(level, item, ms, share, note, fill=None, font=None):
        ws.append([level, item, ms, share, note])
        rr = ws.max_row
        if isinstance(ms, (int, float)):
            ws.cell(rr, 3).number_format = NUM
        if isinstance(share, (int, float)):
            ws.cell(rr, 4).number_format = PCT
        if fill:
            for cc in range(1, 6):
                ws.cell(rr, cc).fill = fill
                ws.cell(rr, cc).font = font or WHT
        return rr

    for config in ("c4", "c128"):
        for phase in ("decode", "prefill"):
            for plat in ("B200", "MI355"):
                r = results.get((config, phase, plat))
                if not r:
                    continue
                a = r["agg"]; bm = bench.get((config, plat), {})
                d = detail.get((config, phase, plat), {})
                steps = a["n"] or 1
                work_ps = a["work"] / 1000
                tot_ms = d["total_us"] / 1000 if d.get("total_us") else None
                rep = max(1, round(tot_ms / work_ps)) if (tot_ms and work_ps) else 1
                drow(f"{config} {phase.upper()}", plat, "", "", "", fill=HF, font=WHT)

                # (1) E2E
                if phase == "decode":
                    e2e = bm.get("e2e_med"); ttft = bm.get("ttft_med"); tp = bm.get("tpot_med")
                    drow("① E2E", "median E2E latency", e2e, None,
                         f"~{IN} in / ~0.8-1.0x{OUT} out tokens", fill=L1)
                    drow("", "├ TTFT median (prefill part)", ttft, None, "time to first token")
                    dec = (e2e - ttft) if (e2e and ttft) else None
                    ntok = round(dec / tp) if (dec and tp) else None
                    drow("", "└ decode part (E2E - TTFT)", dec, None,
                         f"≈ {ntok} tokens × TPOT {tp} ms/token")
                else:
                    drow("① E2E", "median TTFT (prefill latency)", bm.get("ttft_med"), None,
                         "full request: queue + tokenize + chunked forwards", fill=L1)
                    drow("", "mean TTFT", bm.get("ttft_mean"), None, "queue-inflated (esp. c128)")

                # (2) step (trace)
                unit = "per decode step (=1 tok)" if phase == "decode" else "per prefill forward"
                note = (f"≈ ITL median {bm.get('itl_med')}") if phase == "decode" else "one chunked forward"
                drow("② STEP", f"GPU-active union — {unit}", a["union"]/1000, None, note, fill=L2)
                drow("", "kernel-sum (naive)", a["work"]/1000, None, "multi-stream overcount")
                drow("", "step wall", a["wall"]/1000, None, "incl. profiling idle bubbles")
                drow("", "toolbox step windows", steps, None, f"analyzer rep = {rep} forward(s)")

                # (3) layer (analyzer): ms = share% x per-step total (robust across
                # Summary rollups and Module-Tree fallback)
                if tot_ms:
                    per_step = tot_ms/rep
                    drow("③ LAYER", "total attributed / step", per_step, None,
                         "sum of per-module kernel time", fill=L3)
                    for label, pct in d.get("layers", []):
                        ms = (pct/100*per_step) if pct is not None else None
                        drow("", f"  {label}", ms, pct, "")
                else:
                    drow("③ LAYER", "(analyzer detail not available)", None, None, "", fill=L3)

                # (4) category (analyzer), normalised per step by rep
                if d.get("cats"):
                    drow("④ CATEGORY", "top kernel classes / step", None, None, "", fill=L4)
                    for cat, us, pct in d["cats"][:7]:
                        drow("", f"  {cat}", us/1000/rep, pct, "")

                ws.append([])

    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.cell(2, 1).font = Font(italic=True, size=9, color="595959")
    for c in ws[hrow]:
        c.fill = HF; c.font = HFONT; c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(hrow+1, 1).coordinate
    for i, w in enumerate([13, 34, 11, 9, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- Categories ----------------
    ws = wb.create_sheet("Categories")
    ws.append(["Category ms/step by config  (B200 vs MI355, ratio = MI355/B200)"])
    ws.append([])
    top = ["Category"]
    sub = [""]
    blocks = [("c4", "decode"), ("c4", "prefill"), ("c128", "decode"), ("c128", "prefill")]
    for config, phase in blocks:
        top += [f"{config} {phase}", "", ""]
        sub += ["B200", "MI355", "MI355 vs B200 (%)"]
    ws.append(top); ws.append(sub)
    trow = ws.max_row - 1
    # merge top block headers
    col = 2
    for _ in blocks:
        ws.merge_cells(start_row=trow, start_column=col, end_row=trow, end_column=col+2)
        col += 3
    all_cats = set()
    for (config, phase) in blocks:
        for plat in ("B200", "MI355"):
            r = results.get((config, phase, plat))
            if r: all_cats |= set(r["cat"])
    # order categories by total across everything
    def cat_total(c):
        s = 0.0
        for (config, phase) in blocks:
            for plat in ("B200", "MI355"):
                r = results.get((config, phase, plat))
                if r: s += r["cat"].get(c, 0)
        return s
    for c in sorted(all_cats, key=lambda c: -cat_total(c)):
        row = [c]
        for (config, phase) in blocks:
            rb = results.get((config, phase, "B200"))
            rm = results.get((config, phase, "MI355"))
            vb = rb["cat"].get(c, 0)/1000 if rb else 0
            vm = rm["cat"].get(c, 0)/1000 if rm else 0
            ratio = (vm/vb*100) if vb else None
            row += [vb, vm, ratio]
        ws.append(row)
    # totals row (E2E wall & kernel-sum)
    for metric, key in (("GPU-active (ms)", "union"), ("Kernel-sum (ms)", "work"), ("E2E wall (ms)", "wall")):
        row = [metric]
        for (config, phase) in blocks:
            rb = results.get((config, phase, "B200"))
            rm = results.get((config, phase, "MI355"))
            vb = rb["agg"][key]/1000 if rb else 0
            vm = rm["agg"][key]/1000 if rm else 0
            ratio = (vm/vb*100) if vb else None
            row += [vb, vm, ratio]
        ws.append(row)
        for cc in ws[ws.max_row]: cc.font = Font(bold=True)
    ws.cell(1, 1).font = Font(bold=True, size=12)
    for c in ws[trow]:
        c.fill = HF; c.font = HFONT; c.alignment = Alignment(horizontal="center")
    for c in ws[trow+1]:
        c.fill = HF2; c.font = Font(bold=True, color="FFFFFF"); c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B" + str(trow+2)
    ws.column_dimensions["A"].width = 20
    for i in range(2, len(top)+1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    # per-block columns: excel col 2,3 = values (NUM); col 4 = ratio (PCT); repeat every 3
    for r in range(trow+2, ws.max_row+1):
        for col in range(2, len(top)+1):
            fmt = PCT if (col - 1) % 3 == 0 else NUM
            ws.cell(r, col).number_format = fmt

    # ---------------- per-run kernel sheets ----------------
    for (config, phase, plat), r in results.items():
        name = f"{plat}_{config}_{phase}"[:31]
        ws = wb.create_sheet(name)
        ws.append(["exec_order", "category", "kernel_name",
                   "count/step", "total_us/step", "avg_us", "% of work"])
        step_tot = r["agg"]["work"]
        for i, k in enumerate(sorted(r["rows"], key=lambda x: x["first"]), 1):
            ws.append([i, k["category"], k["name"],
                       k["cnt"], k["tps"], k["avg"], k["pct"]])
            for col in (4, 5, 6):
                ws.cell(ws.max_row, col).number_format = NUM
            ws.cell(ws.max_row, 7).number_format = PCT
        style_header(ws)
        for i, w in enumerate([10, 18, 78, 12, 15, 10, 11], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    out = os.path.join(QDIR, "qwen_MI355_vs_B200_report.xlsx")
    wb.save(out)
    print("wrote", out)

if __name__ == "__main__":
    main()
