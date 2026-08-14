"""Decode summary laid out like analysis/example.xlsx.

The template pairs a per-phase wall and kernel table for four configurations
with a per-category kernel table for the phase that dominates the step, and
names the actual kernels behind every category. This script fills that layout
with measured data instead of the placeholder numbers.

    python example_format.py --root <MTP profiles> --mode MTP \
                             --root <non-MTP profiles> --mode nonMTP -o out.xlsx

One sheet per mode. Each sheet holds 8k conc4, 8k conc64, 70k conc4 and
70k conc64, with MI355 paired against B200 at equal running batch.
"""

from __future__ import annotations

import argparse
import os
import sys

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from compare_modes import find_profile_dirs
from mtp_profile.kernels import GLM52_RULES, KernelClassifier
from mtp_profile.phases import phase_order
from mtp_profile.sweep import build_run, discover

BLOCKS = [("8k", 4), ("8k", 64), ("70k", 4), ("70k", 64)]
LABELS = {"8k": "8k", "70k": "70000"}
NUM = "0.00"
MIN_KERNEL_MS = 0.02
MAX_KERNELS_LISTED = 4
MAX_STREAMS_LISTED = 4
WIDTHS = {"A": 11.4, "C": 12.0, "D": 11.6, "G": 21.5, "H": 15.1}


def load(root: str, classifier) -> dict:
    """-> {(context, gpu, variant, conc): run}."""
    runs = {}
    for (ctx, gpu, variant), directory in sorted(find_profile_dirs(root).items()):
        try:
            found = discover(directory)
        except ValueError:
            continue
        for conc in sorted(found):
            runs[(ctx, gpu, variant, conc)] = build_run(gpu, conc, found[conc], classifier)
    return runs


def pick(runs: dict, ctx: str, conc: int):
    """B200 plus the MI355 variant that ran the same batch, if one exists."""
    b200 = runs.get((ctx, "b200", "", conc))
    if b200 is None:
        return None
    candidates = [
        (v, r) for (c, g, v, k), r in runs.items()
        if g == "mi355" and c == ctx and k == conc
    ]
    if not candidates:
        return None
    matched = [(v, r) for v, r in candidates if r.bs == b200.bs]
    variant, mi355 = (matched or sorted(candidates, key=lambda vr: vr[0] or ""))[0]
    note = f"MI355 {variant or 'default'}, bs={mi355.bs}"
    if not matched:
        note += f" vs B200 bs={b200.bs} (batch not matched)"
    return mi355, b200, note


def kernel_list(stats, category: str, classifier) -> str:
    """`name (x ms/step)` for the kernels behind one category, largest first."""
    rows = [
        (k, v) for k, v in stats.per_kernel_ms.items()
        if classifier.classify(k) == category and v >= MIN_KERNEL_MS
    ]
    rows.sort(key=lambda kv: -kv[1])
    return "; ".join(f"{k} ({v:.3f} ms/step)" for k, v in rows[:MAX_KERNELS_LISTED])


def num(ws, row: int, col: int, value):
    """Write a rounded value rather than a formula.

    The template computes the diff and total columns with formulas, but a
    formula written outside Excel carries no cached result, so every reader
    except Excel itself sees a blank cell.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = NUM
    return cell


def write_phase_block(ws, row: int, ctx: str, conc: int, mi355, b200, phases) -> int:
    ws.cell(row=row, column=3, value=f"conc{conc}")
    ws.cell(row=row, column=4, value="Phase")
    for col, text in (
        (5, "B200 "), (6, "MI355"), (7, "Δ"), (8, "B200 "), (9, "MI355"), (10, "Δ")
    ):
        ws.cell(row=row, column=col, value=text)
    row += 1

    wall_b = wall_a = kern_b = kern_a = 0.0
    for phase in phases:
        sa, sb = mi355.stats.get(phase), b200.stats.get(phase)
        wb_, wa = (sb.wall_ms if sb else 0.0), (sa.wall_ms if sa else 0.0)
        kb, ka = (sb.kernel_ms if sb else 0.0), (sa.kernel_ms if sa else 0.0)
        wall_b, wall_a, kern_b, kern_a = wall_b + wb_, wall_a + wa, kern_b + kb, kern_a + ka
        ws.cell(row=row, column=3, value=LABELS[ctx])
        ws.cell(row=row, column=4, value=phase)
        for col, value in (
            (5, wb_), (6, wa), (7, wa - wb_), (8, kb), (9, ka), (10, ka - kb)
        ):
            num(ws, row, col, round(value, 2))
        row += 1

    ws.cell(row=row, column=4, value="Total")
    for col, value in (
        (5, wall_b), (6, wall_a), (7, wall_a - wall_b),
        (8, kern_b), (9, kern_a), (10, kern_a - kern_b),
    ):
        num(ws, row, col, round(value, 2))
    return row + 1


def stream_detail(stats) -> str:
    """`tid N (x ms/step)` for the busiest streams, with the tail summarised.

    B200 spreads a plain decode step over dozens of streams, so listing them
    all would bury the shape of the distribution.
    """
    rows = [kv for kv in sorted(stats.per_stream_ms.items(), key=lambda kv: -kv[1])
            if kv[1] >= 0.01]
    head = "; ".join(f"tid {tid} ({ms:.2f} ms/step)" for tid, ms in rows[:MAX_STREAMS_LISTED])
    tail = rows[MAX_STREAMS_LISTED:]
    if tail:
        head += f"; +{len(tail)} more ({sum(ms for _, ms in tail):.2f} ms/step)"
    return head


def write_stream_block(ws, row: int, ctx: str, conc: int, mi355, b200, phases) -> int:
    """Split the wall gap into kernel work and multi-stream scheduling.

    A phase spread over several streams finishes in less wall time than the
    kernels add up to, so wall minus kernel is negative exactly when streams
    overlap. That term is the whole difference between the two Δ columns of the
    phase table above, and it is what CUDA multi-stream execution buys.
    """
    for start, end, text in (
        (5, 6, "streams"),
        (7, 8, "overlap = kernel/wall"),
        (9, 10, "idle ms/step = wall - kernel"),
        (11, 13, "wall Δ = kernel Δ + idle Δ"),
        (14, 15, "kernel ms/step per stream"),
    ):
        ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        ws.cell(row=row, column=start, value=text)
    row += 1

    ws.cell(row=row, column=3, value=f"conc{conc}")
    ws.cell(row=row, column=4, value="Phase")
    for col, text in (
        (5, "B200 "), (6, "MI355"), (7, "B200 "), (8, "MI355"), (9, "B200 "),
        (10, "MI355"), (11, "Δ idle"), (12, "Δ kernel"), (13, "Δ wall"),
        (14, "B200 streams"), (15, "MI355 streams"),
    ):
        ws.cell(row=row, column=col, value=text)
    row += 1

    wall_b = wall_a = kern_b = kern_a = 0.0
    tids_b, tids_a = set(), set()
    for phase in phases:
        sa, sb = mi355.stats[phase], b200.stats[phase]
        wall_b, wall_a = wall_b + sb.wall_ms, wall_a + sa.wall_ms
        kern_b, kern_a = kern_b + sb.kernel_ms, kern_a + sa.kernel_ms
        tids_b |= set(sb.per_stream_ms)
        tids_a |= set(sa.per_stream_ms)
        idle_b, idle_a = sb.wall_ms - sb.kernel_ms, sa.wall_ms - sa.kernel_ms

        ws.cell(row=row, column=3, value=LABELS[ctx])
        ws.cell(row=row, column=4, value=phase)
        ws.cell(row=row, column=5, value=len(sb.per_stream_ms))
        ws.cell(row=row, column=6, value=len(sa.per_stream_ms))
        for col, value in (
            (7, sb.overlap_factor), (8, sa.overlap_factor),
            (9, idle_b), (10, idle_a),
            (11, idle_a - idle_b),
            (12, sa.kernel_ms - sb.kernel_ms),
            (13, sa.wall_ms - sb.wall_ms),
        ):
            num(ws, row, col, round(value, 2))
        ws.cell(row=row, column=14, value=stream_detail(sb))
        ws.cell(row=row, column=15, value=stream_detail(sa))
        row += 1

    idle_b, idle_a = wall_b - kern_b, wall_a - kern_a
    ws.cell(row=row, column=4, value="Total")
    ws.cell(row=row, column=5, value=len(tids_b))
    ws.cell(row=row, column=6, value=len(tids_a))
    for col, value in (
        (7, kern_b / wall_b if wall_b else 0.0),
        (8, kern_a / wall_a if wall_a else 0.0),
        (9, idle_b), (10, idle_a),
        (11, idle_a - idle_b),
        (12, kern_a - kern_b),
        (13, wall_a - wall_b),
    ):
        num(ws, row, col, round(value, 2))
    return row + 1


def write_category_block(ws, row: int, ctx: str, conc: int, mi355, b200, phase,
                         note: str, classifier) -> int:
    ws.cell(row=row, column=1, value=f"{LABELS[ctx]}-conc{conc}")
    ws.cell(row=row, column=2, value=note)
    row += 1

    ws.cell(row=row, column=1, value=phase)
    for col, text in (
        (2, "Category"),
        (3, "B200 kernel ms/step"),
        (4, "MI355 kernel ms/step"),
        (5, "Δ (MI355-B200)"),
        (6, "B200 kernels"),
        (7, "MI355 kernels"),
    ):
        ws.cell(row=row, column=col, value=text)
    row += 1

    sa, sb = mi355.stats[phase], b200.stats[phase]
    cats = set(sa.per_category_ms) | set(sb.per_category_ms)
    ordered = sorted(
        cats,
        key=lambda c: -(sa.per_category_ms.get(c, 0.0) - sb.per_category_ms.get(c, 0.0)),
    )

    total_a = total_b = 0.0
    for cat in ordered:
        va = sa.per_category_ms.get(cat, 0.0)
        vb = sb.per_category_ms.get(cat, 0.0)
        if max(va, vb) < MIN_KERNEL_MS:
            continue
        total_a += va
        total_b += vb
        ws.cell(row=row, column=1, value=LABELS[ctx])
        ws.cell(row=row, column=2, value=cat)
        num(ws, row, 3, round(vb, 2))
        num(ws, row, 4, round(va, 2))
        num(ws, row, 5, round(va - vb, 2))
        ws.cell(row=row, column=6, value=kernel_list(sb, cat, classifier))
        ws.cell(row=row, column=7, value=kernel_list(sa, cat, classifier))
        row += 1

    row += 1
    num(ws, row, 3, round(total_b, 2))
    num(ws, row, 4, round(total_a, 2))
    num(ws, row, 5, round(total_a - total_b, 2))
    return row + 2


def build_sheet(ws, runs: dict, classifier):
    selected = []
    for ctx, conc in BLOCKS:
        chosen = pick(runs, ctx, conc)
        if chosen is None:
            print(f"  [skip] {ctx} conc{conc}: no matching pair")
            continue
        selected.append((ctx, conc) + chosen)

    phases = phase_order(*[r.stats for _, _, a, b, _ in selected for r in (a, b)])
    # The phase worth breaking into categories: the one holding most of the step.
    main_phase = max(
        phases, key=lambda p: sum(r.stats[p].kernel_ms for _, _, a, b, _ in selected for r in (a, b))
    )

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
    ws.cell(row=2, column=5, value="wall ms/step")
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=10)
    ws.cell(row=2, column=8, value="kernel ms/step")

    row = 3
    for i, (ctx, conc, mi355, b200, _note) in enumerate(selected):
        row = write_phase_block(ws, row, ctx, conc, mi355, b200, phases)
        if i + 1 < len(selected) and selected[i + 1][0] != ctx:
            row += 1
    row += 1

    ws.cell(row=row, column=1, value="multi-stream")
    ws.cell(
        row=row,
        column=2,
        value="how much of the wall gap is scheduling rather than kernel work",
    )
    row += 2
    for i, (ctx, conc, mi355, b200, _note) in enumerate(selected):
        row = write_stream_block(ws, row, ctx, conc, mi355, b200, phases)
        row += 1 if i + 1 < len(selected) and selected[i + 1][0] != ctx else 0
    row += 1

    for ctx, conc, mi355, b200, note in selected:
        row = write_category_block(
            ws, row, ctx, conc, mi355, b200, main_phase, note, classifier
        )
        print(f"  {ctx} conc{conc}: {note}")

    for letter, width in WIDTHS.items():
        ws.column_dimensions[letter].width = width


def main(argv=None):
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--root", action="append", required=True)
    ap.add_argument("--mode", action="append", required=True,
                    help="sheet name for the matching --root")
    ap.add_argument("-o", "--output", required=True, metavar="XLSX")
    ap.add_argument("--rules", default=GLM52_RULES)
    args = ap.parse_args(argv)

    if len(args.root) != len(args.mode):
        raise SystemExit("--root and --mode must be given the same number of times")

    classifier = KernelClassifier(args.rules)
    wb = Workbook()
    wb.remove(wb.active)
    for mode, root in zip(args.mode, args.root):
        print(f"[{mode}]")
        build_sheet(wb.create_sheet(mode), load(root, classifier), classifier)

    out = os.path.abspath(args.output)
    wb.save(out)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
