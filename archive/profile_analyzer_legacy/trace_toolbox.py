#!/usr/bin/env python
"""
Trace Toolbox -- one CLI for GPU-trace (Kineto / Chrome trace) analysis and
prefill/decode splitting. Works across AMD (roctracer) and NVIDIA (CUPTI),
any model, with or without CUDA/HIP graphs.

Subcommands
-----------
  analyze   Analyze decode/prefill windows: per-kernel + category buckets +
            concurrency factor + Excel. Auto-detects phase / batch-size /
            graph-mode / platform.
  compare   Compare N traces side by side (category rollup + timing/concurrency).
  opmap     Build kernel -> op/module/source JSON map from a graph-DISABLED trace
            (feed to `analyze --op-map`).
  split     Split a combined end-to-end trace into standalone prefill + decode
            trace files (merges the old scan/prep/boundary/split steps).
  validate  Validate JSON integrity of trace file(s) (e.g. split output).
  steps     Quick scan of step/run_batch annotations and phase transitions.

Examples
--------
  python trace_toolbox.py analyze run.trace.json --phase decode --out d.xlsx
  python trace_toolbox.py compare a.json b.json --labels MI355,B200 --phase decode
  python trace_toolbox.py opmap run_graph_dis.trace.json map.json
  python trace_toolbox.py split e2e.trace.json          # -> *.prefill/*.decode
  python trace_toolbox.py validate x.prefill.trace.json x.decode.trace.json
  python trace_toolbox.py steps e2e.trace.json
"""
import re, sys, os, json, bisect, time, argparse
from collections import defaultdict, Counter

# ============================================================ shared regexes
CAT_RE   = re.compile(r'"cat":\s*"([a-z_]+)"')
NAME_RE  = re.compile(r'"name":\s*"((?:[^"\\]|\\.)*)"')
NAME_S   = re.compile(r'"name":\s*"([^"]*)"')
TSDUR_RE = re.compile(r'"ts":\s*([0-9.]+),\s*"dur":\s*([0-9.]+)')
TS_RE    = re.compile(r'"ts":\s*([0-9.]+)')
DUR_RE   = re.compile(r'"dur":\s*([0-9.]+)')
TID_RE   = re.compile(r'"tid":\s*(\d+)')
PID_RE   = re.compile(r'"pid":\s*(\d+)')
PIDTID_RE= re.compile(r'"pid":\s*(\d+),\s*"tid":\s*(\d+)')
EXTID_RE = re.compile(r'"External id":\s*(\d+)')
CORR_RE  = re.compile(r'"correlation":\s*(\d+)')
BS_RE    = re.compile(r'bs=(\d+)')

def is_decode_marker(name):
    return name.startswith("step[DECODE") or "TARGET_VERIFY" in name
def is_prefill_marker(name):
    return name.startswith("step[EXTEND")

# ============================================================ classification
CAT_RULES = [
    ("Communication", ["cross_device_reduce", "all_reduce", "allreduce", "all_gather",
                        "allgather", "reduce_scatter", "nccl", "rccl", "mnnvl",
                        "one_shot", "two_shot", "oneshot"]),
    ("MoE",           ["moe_sorting", "opus_moe", "ck_moe", "fused_experts", "moe_gemm",
                        "mfma_moe", "e2m1", "routing", "finalizekernel", "topkgating",
                        "gating", "gate_sigmoid", "sigmoid_mul", "moe", "expert", "permute"]),
    ("Attention",     ["mqa_logits", "paged_mqa", "paged_decode", "paged_prefill", "flash",
                        "fmha", "mla_", "_mla", "attn", "attention", "hadamard",
                        "gated_delta", "causal_conv1d", "unified_attention", "reduce_segments",
                        "store_kvcache", "qkvzba", "set_kv_buffer", "fwd_kernel_stage",
                        "mla_dec"]),
    ("Rotary/Embed",  ["rope", "rotary", "sbhd_cached", "mrope"]),
    ("Norm",          ["rmsnorm", "rms_norm", "layernorm", "layernorm2d", "layer_norm", "_norm"]),
    ("Quant/Scale",   ["quant", "dequant", "scaled", "act_quant", "absmax", "nvfp4",
                        "per_token", "per_group", "per_tensor", "fp8", "int8", "cast", "convert"]),
    ("GEMM/Linear",   ["gemm", "cijk", "cutlass", "hipblas", "rocblas", "nvjet", "wv_split",
                        "blockscale", "tensile", "matmul", "_mm_", "bmm", "addmm", "hgemm",
                        "splitkreduce"]),
    ("Activation",    ["act_and_mul", "silu", "gelu", "swiglu", "glu", "activation"]),
    ("Sampling/TopK", ["topk", "top_k", "argmax", "argsort", "sort", "sample", "greedy",
                        "multinomial", "softmax", "logprob", "penal", "temperature"]),
    ("Embedding",     ["index_select", "indexselect", "embedding", "embed", "vocab"]),
    ("Reduction",     ["rocprim", "cumsum", "scan", "reduce", "trampoline", "lookback",
                        "hipcub", "cub::", "prefix"]),
    ("Elementwise/Memory", ["elementwise", "vectorized", "manual_unroll", "copy", "catarray",
                            "cat", "fill", "index", "clamp", "masked", "arange", "memcpy",
                            "memset", "gather", "scatter", "pad", "contiguous", "stride",
                            "multi_tensor"]),
]

def classify(kname):
    low = kname.lower()
    for cat, pats in CAT_RULES:
        for p in pats:
            if p in low:
                return cat
    return "Other"

# ============================================================ op-map lookups
_IDX_RE = re.compile(r"_\d+$")

def make_map_lookups(raw_map):
    def best_op(k):
        e = raw_map.get(k)
        if not e: return "<no-map>"
        real = [(o, c) for o, c in e.get("op", []) if o != "<unmapped>"]
        return real[0][0] if real else "<unmapped>"
    def best_module(k):
        e = raw_map.get(k)
        if not e: return "<no-map>"
        agg = defaultdict(int)
        for m, c in e.get("module", []):
            if m == "<none>": continue
            agg[_IDX_RE.sub("", m)] += c
        return max(agg.items(), key=lambda kv: kv[1])[0] if agg else "<none>"
    def best_source(k):
        e = raw_map.get(k)
        if not e: return "<no-map>"
        real = [(s, c) for s, c in e.get("source", []) if s != "<none>"]
        return real[0][0] if real else "<none>"
    return best_op, best_module, best_source

# ============================================================ parse (analyze)
def merge_stream_windows(ivs):
    """Collapse time-overlapping step windows into one.

    On multi-stream platforms (e.g. B200 running 2+ CUDA streams), a single
    step's gpu_user_annotation (step[DECODE]/step[EXTEND]) is emitted once per
    stream on a different tid. Those per-stream windows cover (nested/overlapping)
    the SAME step, so counting them all inflates the step count and double-counts
    each kernel (aggregate() collects kernels by time, not tid). Merging
    overlapping intervals folds each step's per-stream windows back into one
    window spanning the widest range, giving the true step count. Sequential
    distinct steps don't overlap, so they are left untouched.
    """
    if not ivs:
        return ivs
    ivs = sorted(ivs)
    merged = [ivs[0]]
    for s, e in ivs[1:]:
        ls, le = merged[-1]
        if s < le:                       # overlaps prev -> same step, other stream
            if e > le:
                merged[-1] = (ls, e)
        else:
            merged.append((s, e))
    return merged

def parse_trace(path, sample_k=None, sample_bs=None, guard_us=2000.0, seek_frac=None,
                sample_phase="decode"):
    """windows{key:[(ts,end)]}, kernels[(ts,name,dur)] (sorted), graph flags, info.

    Sampling (for huge files): if sample_k is set, stop reading as soon as K
    complete decode steps have been captured (a step is 'complete' once kernels
    with ts beyond its end + guard have been seen, so all its kernels are in).
    With sample_bs, target that batch size (recommended for steady state); else
    the first bs to reach K closed steps wins. Decode steps repeat near-identically,
    so K sampled steps are representative of the whole run.
    """
    intern = {}
    windows = defaultdict(list)
    kts=[]; knm=[]; kdur=[]
    pending=None; pkey=None; pname=None
    graph = {"hipGraph": False, "cudaGraph": False, "CompiledFxGraph": False}
    max_kts = 0.0; n_lines = 0; stopped = False; chosen_bs = None
    seek_byte = 0
    with open(path, encoding="utf-8") as f:
        if seek_frac:
            seek_byte = int(os.path.getsize(path) * seek_frac)
            f.seek(seek_byte); f.readline()  # discard partial line after the seek
        for line in f:
            n_lines += 1
            if pending is not None:
                m = TSDUR_RE.search(line)
                if m:
                    ts=float(m.group(1)); dur=float(m.group(2))
                    if pending == "rb":
                        windows[pkey].append((ts, ts+dur))
                        # early-exit check for sampling
                        if sample_k and pkey[0]==sample_phase and (sample_bs is None or pkey[1]==sample_bs):
                            bkey = pkey
                            closed = sum(1 for s,e in windows[bkey] if e + guard_us <= max_kts)
                            if closed >= sample_k:
                                chosen_bs = bkey[1]; stopped = True; pending=None; break
                    else:
                        kts.append(ts); knm.append(pname); kdur.append(dur)
                        if ts > max_kts: max_kts = ts
                pending=None; continue
            if not graph["hipGraph"] and "hipGraphLaunch" in line: graph["hipGraph"]=True
            if not graph["cudaGraph"] and "cudaGraphLaunch" in line: graph["cudaGraph"]=True
            if not graph["CompiledFxGraph"] and "CompiledFxGraph" in line: graph["CompiledFxGraph"]=True
            m = CAT_RE.search(line)
            if not m: continue
            c = m.group(1)
            if c == "gpu_user_annotation":
                nm = NAME_RE.search(line)
                if not nm: continue
                name = nm.group(1)
                if name.startswith("step[DECODE"):
                    b = BS_RE.search(name); pkey = ("decode", int(b.group(1)) if b else 0); pending="rb"
                elif name.startswith("step[EXTEND"):
                    pkey = ("prefill", None); pending="rb"
                elif name == "scheduler.run_batch":
                    pkey = ("run_batch", None); pending="rb"
            elif c == "kernel":
                nm = NAME_RE.search(line)
                if nm: s=nm.group(1); pname=intern.setdefault(s, s); pending="kern"
    order = sorted(range(len(kts)), key=lambda i: kts[i])
    kernels = [(kts[i], knm[i], kdur[i]) for i in order]
    # Fold multi-stream duplicate step windows so each step is counted once.
    for k in list(windows.keys()):
        windows[k] = merge_stream_windows(windows[k])
    info = {"sampled": stopped, "chosen_bs": chosen_bs, "lines_read": n_lines,
            "seek_byte": seek_byte,
            "decode_freq": {k[1]: len(v) for k, v in windows.items() if k[0]=="decode"},
            "prefill_count": len(windows.get(("prefill", None), []))}
    return windows, kernels, graph, info

def aggregate(windows_list, kernels):
    kts = [k[0] for k in kernels]
    n = len(windows_list)
    cnt=defaultdict(int); tot=defaultdict(float); first=defaultdict(lambda: float("inf"))
    durs=defaultdict(list); wsum=bsum=usum=0.0
    for ts, end in windows_list:
        wsum += end - ts
        lo = bisect.bisect_left(kts, ts); hi = bisect.bisect_left(kts, end)
        ivs=[]
        for i in range(lo, hi):
            _, nm, dur = kernels[i]
            cnt[nm]+=1; tot[nm]+=dur; durs[nm].append(dur); bsum+=dur
            if kernels[i][0] < first[nm]: first[nm]=kernels[i][0]
            ivs.append((kernels[i][0], kernels[i][0]+dur))
        ivs.sort(); ce=-1.0
        for s, e in ivs:
            if s > ce: usum += e - s; ce = e
            elif e > ce: usum += e - ce; ce = e
    return dict(n=n, cnt=cnt, tot=tot, durs=durs, first=first,
                wall=wsum/n if n else 0, work=bsum/n if n else 0, union=usum/n if n else 0)

def select_windows(windows, phase, bs):
    decode_keys = sorted([k for k in windows if k[0]=="decode"], key=lambda k:-len(windows[k]))
    have_decode=bool(decode_keys); have_prefill=("prefill",None) in windows; have_rb=("run_batch",None) in windows
    def decode_sel():
        if bs == "all":
            merged=[]; [merged.extend(windows[k]) for k in decode_keys]; return [("decode bs=all", merged)]
        if bs is not None:
            k=("decode", int(bs)); return [(f"decode bs={bs}", windows.get(k, []))]
        k=decode_keys[0]; return [(f"decode bs={k[1]}", windows[k])]
    out=[]
    if phase=="auto":
        if have_decode: out=decode_sel()
        elif have_prefill: out=[("prefill", windows[("prefill",None)])]
        elif have_rb: out=[("run_batch", windows[("run_batch",None)])]
    elif phase=="decode": out=decode_sel()
    elif phase=="prefill": out=[("prefill", windows.get(("prefill",None), []))]
    elif phase=="run_batch": out=[("run_batch", windows.get(("run_batch",None), []))]
    elif phase=="all":
        if have_decode: out+=decode_sel()
        if have_prefill: out+=[("prefill", windows[("prefill",None)])]
        if not have_decode and not have_prefill and have_rb: out+=[("run_batch", windows[("run_batch",None)])]
    return [(lbl, w) for lbl, w in out if w]

def analyze(path, phase="auto", bs=None, op_map=None, sample=None, seek=None):
    sample_phase = "prefill" if phase == "prefill" else "decode"
    sample_bs = int(bs) if (sample and sample_phase == "decode" and bs not in (None, "all")) else None
    windows, kernels, graph, info = parse_trace(path, sample_k=sample, sample_bs=sample_bs,
                                                seek_frac=seek, sample_phase=sample_phase)
    # when sampling decode picked a bs (no --bs given), analyze that bs
    if sample and sample_phase == "decode" and bs in (None, "all") and info.get("chosen_bs") is not None:
        bs = info["chosen_bs"]; phase = "decode"
    raw_map=None
    if op_map and os.path.exists(op_map):
        with open(op_map, encoding="utf-8") as f: raw_map=json.load(f)
    bo, bm, bs_ = make_map_lookups(raw_map) if raw_map else (None,None,None)
    # when sampling, only keep the K complete windows we actually captured
    if sample:
        if sample_phase == "prefill":
            key = ("prefill", None)
        else:
            tb = bs if bs not in (None, "all") else info.get("chosen_bs")
            key = ("decode", int(tb)) if tb is not None else None
        if key in windows:
            maxk = max((k[0] for k in kernels), default=0.0)
            complete = [(s,e) for s,e in windows[key] if e <= maxk]
            windows[key] = complete[:sample]
    sel = select_windows(windows, phase, bs)
    sections=[]
    for lbl, wlist in sel:
        agg = aggregate(wlist, kernels); step_tot=agg["work"]; n=agg["n"]; rows=[]
        for nm in agg["cnt"]:
            c=agg["cnt"][nm]; t=agg["tot"][nm]
            rows.append(dict(name=nm, category=classify(nm),
                             module=bm(nm) if raw_map else "n/a",
                             source=bs_(nm) if raw_map else "n/a",
                             op=bo(nm) if raw_map else "n/a",
                             cnt=c/n, tps=t/n, avg=t/c if c else 0,
                             pct=t/(step_tot*n)*100 if step_tot else 0, first=agg["first"][nm]))
        sections.append((lbl, agg, rows))
    return windows, graph, raw_map is not None, sections, info

# ============================================================ Excel (analyze)
def write_excel(out_path, sections, graph, opmap_used, trace_name):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    hf=PatternFill("solid", fgColor="305496"); hfont=Font(color="FFFFFF", bold=True); PCT='0.00"%"'
    def hdr(ws):
        for c in ws[1]: c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center")
        ws.freeze_panes="A2"
    ws0=wb.create_sheet("summary")
    ws0.append(["Trace", trace_name])
    ws0.append(["Graph mode", ", ".join([k for k,v in graph.items() if v]) or "none (eager)"])
    ws0.append(["op-map", "used" if opmap_used else "not supplied (name-based only)"])
    ws0.append([])
    ws0.append(["Phase","Steps","Wall ms/step","GPU work ms/step","GPU-active ms/step","Concurrency x","Kernels/step"])
    for lbl, agg, _r in sections:
        conc=agg["work"]/agg["union"] if agg["union"] else 0
        ws0.append([lbl, agg["n"], round(agg["wall"]/1000,3), round(agg["work"]/1000,3),
                    round(agg["union"]/1000,3), round(conc,2),
                    round(sum(agg["cnt"].values())/agg["n"],0) if agg["n"] else 0])
    for c in ws0[5]: c.fill=hf; c.font=hfont
    for i,w in enumerate([16,10,16,18,18,14,14],1): ws0.column_dimensions[get_column_letter(i)].width=w
    for r in (1,2,3): ws0.cell(r,1).font=Font(bold=True)
    for lbl, agg, rows in sections:
        n=agg["n"]; step_tot=agg["work"]; safe=re.sub(r'[^A-Za-z0-9=]+','_',lbl)[:24]
        ws=wb.create_sheet(safe[:31])
        ws.append(["exec_order","category","module","source","op","kernel_name",
                   "count/step","total_us/step","avg_us","% of work"])
        for i,r in enumerate(sorted(rows, key=lambda r:r["first"]),1):
            ws.append([i, r["category"], r["module"], r["source"], r["op"], r["name"],
                       round(r["cnt"],2), round(r["tps"],2), round(r["avg"],2), round(r["pct"],2)])
            ws.cell(ws.max_row,10).number_format=PCT
        hdr(ws)
        for i,w in enumerate([10,16,22,38,28,72,12,15,10,11],1): ws.column_dimensions[get_column_letter(i)].width=w
        wc=wb.create_sheet(f"{safe}_cat"[:31]); cagg=defaultdict(lambda:[0.0,0.0])
        for r in rows: cagg[r["category"]][0]+=r["cnt"]; cagg[r["category"]][1]+=r["tps"]
        wc.append(["category","count/step","total_us/step","ms/step","% of work"])
        for cat in sorted(cagg,key=lambda c:-cagg[c][1]):
            cc,tt=cagg[cat]; wc.append([cat,round(cc,2),round(tt,2),round(tt/1000,3),
                       round(tt/step_tot*100,2) if step_tot else 0]); wc.cell(wc.max_row,5).number_format=PCT
        wc.append(["TOTAL",round(sum(v[0] for v in cagg.values()),2),round(step_tot,2),
                   round(step_tot/1000,3),100.0]); wc.cell(wc.max_row,5).number_format=PCT
        for c in wc[wc.max_row]: c.font=Font(bold=True)
        hdr(wc)
        for i,w in enumerate([20,14,16,12,11],1): wc.column_dimensions[get_column_letter(i)].width=w
        if opmap_used:
            wm=wb.create_sheet(f"{safe}_mod"[:31]); magg=defaultdict(lambda:[0.0,0.0]); fbm={}
            for r in sorted(rows,key=lambda r:r["first"]):
                magg[r["module"]][0]+=r["cnt"]; magg[r["module"]][1]+=r["tps"]; fbm.setdefault(r["module"], r["first"])
            wm.append(["module","count/step","total_us/step","% of work"])
            for m in sorted(magg,key=lambda m:fbm[m]):
                cc,tt=magg[m]; wm.append([m,round(cc,2),round(tt,2),round(tt/step_tot*100,2) if step_tot else 0])
                wm.cell(wm.max_row,4).number_format=PCT
            hdr(wm)
            for i,w in enumerate([26,14,16,11],1): wm.column_dimensions[get_column_letter(i)].width=w
    wb.save(out_path)

# ============================================================ opmap builder
def build_op_map(path, out):
    SRC_KEEP="sglang/srt/"
    SRC_EXCLUDE=("/layers/linear.py","/layers/quantization/","fp8_utils.py","layers/utils/multi_platform.py")
    src_re=re.compile(r'([^/]+\.py)\((\d+)\):\s*(.*)$')
    def src_label(nm):
        if SRC_KEEP not in nm: return None
        for ex in SRC_EXCLUDE:
            if ex in nm: return None
        m=src_re.search(nm)
        if not m: return None
        return f"{m.group(1)}:{m.group(3).strip() or '?'}"
    extid_to_op={}; corr_to_extid={}; corr_to_launch={}
    mods_by_tid=defaultdict(list); srcs_by_tid=defaultdict(list); kernels=[]
    cur=name=tid=ts=dur=extid=corr=None
    def commit():
        if cur=="cpu_op":
            if extid is not None and name is not None: extid_to_op[extid]=name
        elif cur=="cuda_runtime":
            if corr is not None:
                if extid is not None: corr_to_extid[corr]=extid
                if ts is not None and tid is not None: corr_to_launch[corr]=(ts,tid)
        elif cur=="python_function":
            if name is not None and ts is not None and dur is not None and tid is not None:
                if name.startswith("nn.Module:"):
                    mods_by_tid[tid].append((ts,ts+dur,name[len("nn.Module:"):].strip()))
                else:
                    lbl=src_label(name)
                    if lbl is not None: srcs_by_tid[tid].append((ts,ts+dur,lbl))
        elif cur=="kernel":
            if name is not None: kernels.append((name,extid,corr))
    with open(path, encoding="utf-8") as f:
        for line in f:
            if '"ph":' in line and '"cat":' in line:
                commit()
                m=CAT_RE.search(line); cur=m.group(1) if m else None
                nm=NAME_RE.search(line); name=nm.group(1) if nm else None
                tm=TID_RE.search(line); tid=int(tm.group(1)) if tm else None
                ts=dur=extid=corr=None
            elif cur in ("cpu_op","cuda_runtime","kernel","python_function"):
                if ts is None and '"ts"' in line:
                    mm=TS_RE.search(line); ts=float(mm.group(1)) if mm else None
                if dur is None and '"dur"' in line:
                    mm=DUR_RE.search(line); dur=float(mm.group(1)) if mm else None
                if extid is None and "External id" in line:
                    mm=EXTID_RE.search(line); extid=int(mm.group(1)) if mm else None
                if corr is None and "correlation" in line:
                    mm=CORR_RE.search(line); corr=int(mm.group(1)) if mm else None
        commit()
    launch_by_tid=defaultdict(list)
    for c,(lts,ltid) in corr_to_launch.items(): launch_by_tid[ltid].append((lts,c))
    def nest_resolve(iv_by_tid):
        out={}
        for t,ivs in iv_by_tid.items():
            ivs.sort(key=lambda o:o[0]); launches=sorted(launch_by_tid.get(t,[])); stack=[]; oi=0; nn=len(ivs)
            for lts,c in launches:
                while oi<nn and ivs[oi][0]<=lts: stack.append((ivs[oi][1],ivs[oi][2])); oi+=1
                while stack and stack[-1][0]<lts: stack.pop()
                if stack: out[c]=stack[-1][1]
        return out
    corr_to_mod=nest_resolve(mods_by_tid); corr_to_src=nest_resolve(srcs_by_tid)
    km_op=defaultdict(Counter); km_mod=defaultdict(Counter); km_src=defaultdict(Counter)
    r_op=r_mod=r_src=0
    for kn,ke,kc in kernels:
        op=None
        if ke is not None and ke in extid_to_op: op=extid_to_op[ke]
        elif kc is not None and kc in corr_to_extid and corr_to_extid[kc] in extid_to_op: op=extid_to_op[corr_to_extid[kc]]
        if op is None: op="<unmapped>"
        else: r_op+=1
        km_op[kn][op]+=1
        mod=corr_to_mod.get(kc) if kc is not None else None
        if mod is None: mod="<none>"
        else: r_mod+=1
        km_mod[kn][mod]+=1
        src=corr_to_src.get(kc) if kc is not None else None
        if src is None: src="<none>"
        else: r_src+=1
        km_src[kn][src]+=1
    res={k:{"op":km_op[k].most_common(),"module":km_mod[k].most_common(),
            "source":km_src[k].most_common()} for k in km_op}
    with open(out,"w",encoding="utf-8") as f: json.dump(res,f,ensure_ascii=False,indent=1)
    tot=max(len(kernels),1)
    print(f"kernels={len(kernels)} distinct={len(km_op)}  resolved op={r_op/tot*100:.1f}% "
          f"module={r_mod/tot*100:.1f}% source={r_src/tot*100:.1f}%")
    print(f"wrote {out}")

# ============================================================ split pipeline
def compute_boundaries(path):
    """Return (cpu_pid, t_cpu, t_gpu): start ts of the first DECODE run_batch per
    timeline domain. Decode run_batch = one containing a DECODE/TARGET_VERIFY step."""
    run_batch={"cpu":[], "gpu":[]}; steps={"cpu":[], "gpu":[]}; cpu_pids=set()
    pending=None
    with open(path, encoding="utf-8") as f:
        for line in f:
            if pending is not None:
                m=TSDUR_RE.search(line)
                if m:
                    ts=float(m.group(1)); dur=float(m.group(2)); kind,domain,phase=pending
                    if kind=="rb": run_batch[domain].append((ts,ts+dur))
                    else: steps[domain].append((ts,ts+dur,phase))
                pending=None; continue
            if '"cat":' not in line: continue
            cm=CAT_RE.search(line)
            if not cm: continue
            cat=cm.group(1)
            if cat=="user_annotation": domain="cpu"
            elif cat=="gpu_user_annotation": domain="gpu"
            else: continue
            nm=NAME_S.search(line)
            if not nm: continue
            name=nm.group(1)
            if domain=="cpu":
                pm=PID_RE.search(line)
                if pm: cpu_pids.add(int(pm.group(1)))
            if name=="scheduler.run_batch": pending=("rb",domain,None)
            elif name.startswith("step["):
                if is_prefill_marker(name): pending=("step",domain,"prefill")
                elif is_decode_marker(name): pending=("step",domain,"decode")
    cpu_pid = max(cpu_pids) if cpu_pids else None
    res={}
    for domain in ("cpu","gpu"):
        rb=sorted(run_batch[domain]); rb_starts=[s for s,e in rb]; rb_ends=[e for s,e in rb]
        n=len(rb); rb_has=[set() for _ in range(n)]; orphans=[]
        for ts,end,phase in steps[domain]:
            i=bisect.bisect_right(rb_starts,ts)-1
            if i>=0 and rb_ends[i]>=ts: rb_has[i].add(phase)
            else: orphans.append((ts,phase))
        dec_starts=[rb_starts[i] for i in range(n) if "decode" in rb_has[i]]
        dec_starts+=[ts for ts,phase in orphans if phase=="decode"]
        res[domain]=min(dec_starts) if dec_starts else None
    return cpu_pid, res.get("cpu"), res.get("gpu")

_STRIP_STR=re.compile(r'"(?:[^"\\]|\\.)*"')
def _brace_delta(line):
    if '{' not in line and '}' not in line: return 0
    s=_STRIP_STR.sub('', line); return s.count('{')-s.count('}')
_PH_RE=re.compile(r'"ph":\s*"([^"]*)"')

def split_trace(src, cpu_pid, t_cpu, t_gpu):
    def classify_elem(txt):
        mph=_PH_RE.search(txt)
        if mph and mph.group(1)=="M": return True, True
        mts=TS_RE.search(txt)
        if not mts: return True, True
        ts=float(mts.group(1)); mpid=PID_RE.search(txt); pid=int(mpid.group(1)) if mpid else -1
        boundary=t_cpu if pid==cpu_pid else t_gpu
        return (ts<boundary, False) if ts<boundary else (False, True)
    base=src[:-len(".trace.json")] if src.endswith(".trace.json") else os.path.splitext(src)[0]
    out_pre=base+".prefill.trace.json"; out_dec=base+".decode.trace.json"
    t0=time.time(); n_evt=n_pre=n_dec=0; depth=0; in_array=False; buf=[]; wp=wd=False
    with open(src, encoding="utf-8") as f, \
         open(out_pre,"w",encoding="utf-8",newline="") as fp, \
         open(out_dec,"w",encoding="utf-8",newline="") as fd:
        for line in f:
            if not in_array:
                fp.write(line); fd.write(line)
                if '"traceEvents"' in line and line.rstrip().endswith('['): in_array=True
                continue
            stripped=line.lstrip()
            if depth==0 and stripped.startswith(']'):
                trailer=line
                for rest in f: trailer+=rest
                fp.write("\n"+trailer); fd.write("\n"+trailer); break
            buf.append(line); depth+=_brace_delta(line)
            if depth==0 and buf:
                core="".join(buf).rstrip(); buf=[]
                if core.endswith(','): core=core[:-1]
                to_pre,to_dec=classify_elem(core); n_evt+=1
                if to_pre:
                    if wp: fp.write(",\n")
                    fp.write(core); wp=True; n_pre+=1
                if to_dec:
                    if wd: fd.write(",\n")
                    fd.write(core); wd=True; n_dec+=1
    print(f"split done in {time.time()-t0:.0f}s  events={n_evt:,} prefill={n_pre:,} decode={n_dec:,}")
    print(f"  -> {out_pre}\n  -> {out_dec}")
    return out_pre, out_dec

# ============================================================ validate
def validate(path):
    header=[]; trailer=[]; depth=0; in_array=False; buf=[]; n_evt=0; bad=0
    with open(path, encoding="utf-8") as f:
        for line in f:
            if not in_array:
                header.append(line)
                if '"traceEvents"' in line and line.rstrip().endswith('['): in_array=True
                continue
            stripped=line.lstrip()
            if depth==0 and stripped.startswith(']'):
                trailer.append(line); [trailer.append(r) for r in f]; break
            buf.append(line); depth+=_brace_delta(line)
            if depth==0 and buf:
                core="".join(buf).rstrip()
                if core.endswith(','): core=core[:-1]
                buf=[]
                try: json.loads(core)
                except Exception as e:
                    bad+=1
                    if bad<=3: print("  BAD ELEMENT:", repr(core[:160]), "->", e)
                n_evt+=1
    try:
        obj=json.loads("".join(header)+"".join(trailer)); wrapper_ok=True; keys=list(obj.keys())
    except Exception as e:
        wrapper_ok=False; keys=str(e)
    print(f"{os.path.basename(path)}: events={n_evt:,} bad={bad} wrapper_ok={wrapper_ok} keys={keys}")
    return bad==0 and wrapper_ok

# ============================================================ steps scan
def scan_steps(path):
    cats={"cpu":Counter(), "gpu":Counter()}; seq={"cpu":[], "gpu":[]}; pending=None
    with open(path, encoding="utf-8") as f:
        for line in f:
            if pending is not None:
                m=TSDUR_RE.search(line)
                if m:
                    ts=float(m.group(1)); domain,phase=pending; seq[domain].append((ts,phase))
                pending=None; continue
            if '"cat":' not in line: continue
            cm=CAT_RE.search(line)
            if not cm: continue
            cat=cm.group(1)
            if cat=="user_annotation": domain="cpu"
            elif cat=="gpu_user_annotation": domain="gpu"
            else: continue
            nm=NAME_S.search(line)
            if not nm: continue
            name=nm.group(1)
            if name.startswith("step["):
                cats[domain][name.split()[0].replace("step[","").rstrip("]")]+=1
                phase="prefill" if is_prefill_marker(name) else ("decode" if is_decode_marker(name) else "?")
                pending=(domain,phase)
            elif name=="scheduler.run_batch":
                cats[domain]["run_batch"]+=1
    for domain in ("cpu","gpu"):
        print(f"\n[{domain}] markers: {dict(cats[domain])}")
        s=sorted(seq[domain]); trans=[]; last=None
        for ts,ph in s:
            if ph!=last: trans.append((ph,ts)); last=ph
        print(f"  phase transitions ({len(trans)}):", " ".join(f"{p}@{t:.0f}" for p,t in trans[:12]),
              "..." if len(trans)>12 else "")

# ============================================================ CLI
def cmd_analyze(a):
    windows, graph, opmap_used, sections, info = analyze(a.trace, a.phase, a.bs, a.op_map, a.sample, a.seek)
    print(f"trace: {os.path.basename(a.trace)}")
    print(f"graph mode: {', '.join([k for k,v in graph.items() if v]) or 'none (eager)'}   "
          f"op-map: {'yes' if opmap_used else 'no (name-based only)'}")
    if a.sample or a.seek:
        note = "SAMPLED (early-exit)" if info.get("sampled") else "sample not reached / whole region read"
        seekmsg = f" from byte {info['seek_byte']:,}" if info.get("seek_byte") else ""
        print(f"** {note}{seekmsg}: read {info['lines_read']:,} lines; "
              f"target bs={info.get('chosen_bs')}; decode-step freq seen: {info['decode_freq']}")
    print("annotation windows:", {k:len(v) for k,v in windows.items()})
    if not sections: print("!! no matching windows for requested phase"); return
    for lbl, agg, rows in sections:
        conc=agg["work"]/agg["union"] if agg["union"] else 0
        print(f"\n=== {lbl} ===")
        print(f"steps={agg['n']}  wall={agg['wall']/1000:.3f}ms  work_sum={agg['work']/1000:.3f}ms  "
              f"active={agg['union']/1000:.3f}ms  concurrency={conc:.2f}x  "
              f"kernels/step={sum(agg['cnt'].values())/agg['n']:.0f}")
        cagg=defaultdict(float)
        for r in rows: cagg[r["category"]]+=r["tps"]
        for cat in sorted(cagg,key=lambda c:-cagg[c]):
            print(f"   {cat:20s} {cagg[cat]/1000:7.3f} ms  ({cagg[cat]/agg['work']*100:5.1f}%)")
    out=a.out or (os.path.splitext(a.trace)[0]+"_analysis.xlsx")
    write_excel(out, sections, graph, opmap_used, os.path.basename(a.trace))
    print(f"\nwrote {out}")

def cmd_compare(a):
    labels=a.labels.split(",") if a.labels else [os.path.basename(t).split("-")[0] for t in a.traces]
    results=[]
    for lbl,tr in zip(labels,a.traces):
        _w,_g,_o,sections,_i=analyze(tr,a.phase,a.bs,None,a.sample,a.seek)
        if not sections: print(f"[skip] {lbl}: no windows"); continue
        seclbl,agg,rows=sections[0]; cat=defaultdict(float)
        for r in rows: cat[r["category"]]+=r["tps"]
        results.append((lbl,seclbl,agg,cat))
        conc=agg["work"]/agg["union"] if agg["union"] else 0
        print(f"{lbl:16s} [{seclbl}] steps={agg['n']} wall={agg['wall']/1000:.3f}ms "
              f"work={agg['work']/1000:.3f}ms active={agg['union']/1000:.3f}ms conc={conc:.2f}x")
    if not results: print("nothing to compare"); return
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="category_compare"
    hf=PatternFill("solid",fgColor="305496"); hfont=Font(color="FFFFFF",bold=True)
    cats=set()
    for *_x,cat in results: cats|=set(cat)
    base=results[0][3]
    header=["Category"]+[f"{l} ms/step" for l,*_ in results]
    if len(results)>1: header+=[f"{results[i][0]}/{results[0][0]} %" for i in range(1,len(results))]
    ws.append(header)
    for c in sorted(cats,key=lambda c:-base.get(c,0)):
        row=[c]+[round(res[3].get(c,0)/1000,3) for res in results]
        if len(results)>1:
            a0=base.get(c,0)
            for i in range(1,len(results)):
                bi=results[i][3].get(c,0); row.append(round(bi/a0*100,1) if a0 else None)
        ws.append(row)
    def mrow(title,fn):
        row=[title]+[round(fn(res)/1000,3) for res in results]
        if len(results)>1:
            a0=fn(results[0])
            for i in range(1,len(results)): row.append(round(fn(results[i])/a0*100,1) if a0 else None)
        ws.append(row)
        for c in ws[ws.max_row]: c.font=Font(bold=True)
    ws.append([]); mrow("GPU work sum",lambda r:r[2]["work"])
    mrow("GPU-active (union)",lambda r:r[2]["union"]); mrow("Wall-clock/step",lambda r:r[2]["wall"])
    ws.append(["Concurrency x"]+[round(r[2]["work"]/r[2]["union"],2) if r[2]["union"] else 0 for r in results])
    for c in ws[ws.max_row]: c.font=Font(bold=True)
    for c in ws[1]: c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center")
    ws.freeze_panes="B2"; ws.column_dimensions["A"].width=22
    for i in range(2,len(header)+1): ws.column_dimensions[get_column_letter(i)].width=16
    out=a.out or "compare.xlsx"; wb.save(out); print(f"\nwrote {out}")

def cmd_opmap(a): build_op_map(a.dis_trace, a.out)

def cmd_split(a):
    cpu_pid=a.cpu_pid; t_cpu=a.t_cpu; t_gpu=a.t_gpu
    if cpu_pid is None or t_cpu is None or t_gpu is None:
        cpu_pid,t_cpu,t_gpu=compute_boundaries(a.trace)
        print(f"auto boundaries: cpu_pid={cpu_pid} t_cpu={t_cpu} t_gpu={t_gpu}")
    if cpu_pid is None or t_cpu is None or t_gpu is None:
        print("!! could not determine boundaries (no decode run_batch found). "
              "Is this a two-phase (prefill->decode) trace?"); return
    op,od=split_trace(a.trace, int(cpu_pid), float(t_cpu), float(t_gpu))
    if a.validate:
        print("validating..."); validate(op); validate(od)

def cmd_validate(a):
    ok=True
    for p in a.traces: ok=validate(p) and ok
    print("ALL OK" if ok else "VALIDATION FAILED")

def cmd_steps(a): scan_steps(a.trace)

def build_parser():
    ap=argparse.ArgumentParser(prog="trace_toolbox", description="GPU trace analysis + splitting toolbox")
    sub=ap.add_subparsers(dest="cmd", required=True)
    p=sub.add_parser("analyze", help="analyze decode/prefill windows -> Excel")
    p.add_argument("trace"); p.add_argument("--phase", default="auto",
                   choices=["auto","decode","prefill","run_batch","all"])
    p.add_argument("--bs", default=None); p.add_argument("--op-map", dest="op_map", default=None)
    p.add_argument("--sample", type=int, default=None,
                   help="huge files: stop after N complete decode steps (repeats are near-identical). "
                        "Pair with --bs to target steady-state batch size.")
    p.add_argument("--seek", type=float, default=None,
                   help="huge files: jump to this fraction (0..1) of the file before parsing, "
                        "to land in the steady-state region. Combine with --sample.")
    p.add_argument("--out", default=None); p.set_defaults(func=cmd_analyze)
    p=sub.add_parser("compare", help="compare N traces side by side")
    p.add_argument("traces", nargs="+"); p.add_argument("--phase", default="auto")
    p.add_argument("--bs", default=None); p.add_argument("--labels", default=None)
    p.add_argument("--sample", type=int, default=None,
                   help="huge files: stop each trace after N complete decode steps")
    p.add_argument("--seek", type=float, default=None,
                   help="huge files: jump to this fraction (0..1) before parsing")
    p.add_argument("--out", default=None); p.set_defaults(func=cmd_compare)
    p=sub.add_parser("opmap", help="build kernel->op/module/source map from graph-DISABLED trace")
    p.add_argument("dis_trace"); p.add_argument("out"); p.set_defaults(func=cmd_opmap)
    p=sub.add_parser("split", help="split combined e2e trace into prefill+decode files")
    p.add_argument("trace"); p.add_argument("--cpu-pid", dest="cpu_pid", type=int, default=None)
    p.add_argument("--t-cpu", dest="t_cpu", type=float, default=None)
    p.add_argument("--t-gpu", dest="t_gpu", type=float, default=None)
    p.add_argument("--validate", action="store_true", help="validate output files after split")
    p.set_defaults(func=cmd_split)
    p=sub.add_parser("validate", help="validate JSON integrity of trace file(s)")
    p.add_argument("traces", nargs="+"); p.set_defaults(func=cmd_validate)
    p=sub.add_parser("steps", help="scan step/run_batch annotations + phase transitions")
    p.add_argument("trace"); p.set_defaults(func=cmd_steps)
    return ap

def main():
    ap=build_parser(); a=ap.parse_args(); a.func(a)

if __name__ == "__main__":
    main()
